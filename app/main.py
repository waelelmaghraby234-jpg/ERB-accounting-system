from __future__ import annotations

import calendar
import hashlib
import hmac
import io
import json
import os
import secrets
from contextlib import asynccontextmanager
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Annotated, Any, Literal
from uuid import UUID

import jwt
from fastapi import Depends, FastAPI, Header, HTTPException, Query, Request, status
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import BaseModel, EmailStr, Field
from psycopg import Connection
from psycopg.rows import dict_row
from psycopg_pool import ConnectionPool

DATABASE_URL = os.environ.get("DATABASE_URL", "")
JWT_SECRET = os.environ.get("JWT_SECRET", "")
AUTO_MIGRATE = os.environ.get("AUTO_MIGRATE", "true").lower() == "true"
ALGORITHM = "HS256"
MONEY = Decimal("0.0001")

if not DATABASE_URL:
    raise RuntimeError("DATABASE_URL is required")
if len(JWT_SECRET) < 32:
    raise RuntimeError("JWT_SECRET must be at least 32 characters")

pool = ConnectionPool(
    conninfo=DATABASE_URL,
    min_size=1,
    max_size=int(os.environ.get("DB_POOL_MAX", "10")),
    kwargs={"row_factory": dict_row},
    open=False,
)


def money(value: Decimal | float | int | str) -> Decimal:
    return Decimal(str(value)).quantize(MONEY, rounding=ROUND_HALF_UP)


def hash_password(password: str, salt: bytes | None = None) -> str:
    salt = salt or secrets.token_bytes(16)
    digest = hashlib.pbkdf2_hmac("sha256", password.encode(), salt, 310_000)
    return f"pbkdf2_sha256$310000${salt.hex()}${digest.hex()}"


def verify_password(password: str, stored: str) -> bool:
    try:
        algorithm, iterations, salt_hex, digest_hex = stored.split("$", 3)
        if algorithm != "pbkdf2_sha256":
            return False
        digest = hashlib.pbkdf2_hmac(
            "sha256", password.encode(), bytes.fromhex(salt_hex), int(iterations)
        )
        return hmac.compare_digest(digest.hex(), digest_hex)
    except (ValueError, TypeError):
        return False


def create_token(user: dict[str, Any]) -> str:
    now = datetime.now(timezone.utc)
    payload = {
        "sub": str(user["user_id"]),
        "group_id": str(user["group_id"]),
        "company_id": str(user["company_id"]) if user.get("company_id") else None,
        "is_group_admin": bool(user.get("is_group_admin")),
        "role_code": user.get("role_code", "GROUP_ADMIN"),
        "email": user["email"],
        "iat": now,
        "exp": now + timedelta(hours=12),
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=ALGORITHM)


def audit(
    conn: Connection,
    user: dict[str, Any],
    action: str,
    entity_type: str,
    entity_id: Any = None,
    company_id: UUID | None = None,
    details: dict[str, Any] | None = None,
) -> None:
    conn.execute(
        """
        INSERT INTO erp.audit_log
            (group_id, company_id, user_id, action, entity_type, entity_id, details)
        VALUES (%s,%s,%s,%s,%s,%s,%s::jsonb)
        """,
        (
            user["group_id"], company_id, user["user_id"], action,
            entity_type, str(entity_id) if entity_id else None,
            json.dumps(details or {}, ensure_ascii=False, default=str),
        ),
    )


def ensure_account(conn: Connection, group_id: UUID, company_id: UUID, account_id: UUID) -> dict[str, Any]:
    row = conn.execute(
        """
        SELECT a.account_id, a.local_account_code, a.local_account_name,
               ga.account_class, ga.is_intercompany, ga.intercompany_role
        FROM erp.accounts a
        JOIN erp.group_accounts ga
          ON ga.group_account_id=a.group_account_id AND ga.group_id=a.group_id
        WHERE a.group_id=%s AND a.company_id=%s AND a.account_id=%s AND a.is_active=TRUE
        """,
        (group_id, company_id, account_id),
    ).fetchone()
    if not row:
        raise HTTPException(status_code=422, detail="الحساب لا يتبع الشركة المختارة")
    return row


def ensure_open_period(conn: Connection, company_id: UUID, posting_date: date) -> None:
    conn.execute("SELECT erp.ensure_open_period(%s,%s)", (company_id, posting_date))


def create_voucher_db(
    conn: Connection,
    *,
    user: dict[str, Any],
    company_id: UUID,
    voucher_no: str,
    document_date: date,
    posting_date: date,
    description: str | None,
    entries: list[dict[str, Any]],
    source_module: str = "GL",
    external_reference: str | None = None,
    post_immediately: bool = True,
) -> dict[str, Any]:
    ensure_open_period(conn, company_id, posting_date)
    debit = money(sum(money(e.get("debit_amount", 0)) for e in entries))
    credit = money(sum(money(e.get("credit_amount", 0)) for e in entries))
    if debit <= 0 or debit != credit:
        raise HTTPException(status_code=422, detail=f"القيد غير متوازن: مدين {debit} / دائن {credit}")
    if len(entries) < 2:
        raise HTTPException(status_code=422, detail="القيد يحتاج سطرين على الأقل")

    voucher = conn.execute(
        """
        INSERT INTO erp.journal_vouchers
            (group_id, company_id, voucher_no, voucher_type, status,
             document_date, posting_date, description, source_module,
             external_reference, created_by)
        VALUES (%s,%s,%s,'GENERAL','DRAFT',%s,%s,%s,%s,%s,%s)
        RETURNING voucher_id, voucher_no, status, posting_date
        """,
        (
            user["group_id"], company_id, voucher_no, document_date, posting_date,
            description, source_module, external_reference, user["user_id"],
        ),
    ).fetchone()

    for line_no, entry in enumerate(entries, 1):
        debit_amount = money(entry.get("debit_amount", 0))
        credit_amount = money(entry.get("credit_amount", 0))
        if (debit_amount > 0) == (credit_amount > 0):
            raise HTTPException(status_code=422, detail="كل سطر يجب أن يكون مديناً أو دائناً فقط")
        account = ensure_account(conn, user["group_id"], company_id, UUID(str(entry["account_id"])))
        counterparty = entry.get("counterparty_company_id")
        ic_ref = entry.get("intercompany_reference")
        if account["is_intercompany"] and not counterparty:
            raise HTTPException(status_code=422, detail="الشركة المقابلة مطلوبة لحساب Intercompany")
        if account["is_intercompany"] and not ic_ref:
            raise HTTPException(status_code=422, detail="مرجع المعاملة المتبادلة مطلوب")
        conn.execute(
            """
            INSERT INTO erp.journal_entries
                (voucher_id, group_id, company_id, account_id, line_no,
                 entry_description, debit_amount, credit_amount,
                 counterparty_company_id, intercompany_reference,
                 branch_id, cost_center_id)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """,
            (
                voucher["voucher_id"], user["group_id"], company_id,
                entry["account_id"], line_no, entry.get("description"),
                debit_amount, credit_amount, counterparty, ic_ref,
                entry.get("branch_id"), entry.get("cost_center_id"),
            ),
        )
    if post_immediately:
        conn.execute("SELECT erp.post_voucher(%s,%s)", (voucher["voucher_id"], user["user_id"]))
        voucher["status"] = "POSTED"
    return voucher


def seed_default_chart(conn: Connection, group_id: UUID) -> None:
    chart = [
        ("100000", "الأصول", "ASSET", "DEBIT", None, False, False, "NONE"),
        ("110000", "الأصول المتداولة", "ASSET", "DEBIT", "100000", False, False, "NONE"),
        ("111000", "الصندوق", "ASSET", "DEBIT", "110000", True, False, "NONE"),
        ("112000", "البنوك", "ASSET", "DEBIT", "110000", True, False, "NONE"),
        ("113000", "العملاء", "ASSET", "DEBIT", "110000", True, False, "NONE"),
        ("114000", "ضريبة قيمة مضافة مدخلات", "ASSET", "DEBIT", "110000", True, False, "NONE"),
        ("115000", "مستحق من شركات المجموعة", "ASSET", "DEBIT", "110000", True, True, "DUE_FROM"),
        ("120000", "الأصول غير المتداولة", "ASSET", "DEBIT", "100000", False, False, "NONE"),
        ("121000", "الأراضي والمباني", "ASSET", "DEBIT", "120000", True, False, "NONE"),
        ("122000", "الأثاث والتجهيزات", "ASSET", "DEBIT", "120000", True, False, "NONE"),
        ("123000", "السيارات", "ASSET", "DEBIT", "120000", True, False, "NONE"),
        ("124000", "مجمع إهلاك الأصول", "ASSET", "CREDIT", "120000", True, False, "NONE"),
        ("125000", "استثمارات في شركات تابعة", "ASSET", "DEBIT", "120000", True, True, "IC_INVESTMENT"),
        ("200000", "الالتزامات", "LIABILITY", "CREDIT", None, False, False, "NONE"),
        ("210000", "الالتزامات المتداولة", "LIABILITY", "CREDIT", "200000", False, False, "NONE"),
        ("211000", "الموردون", "LIABILITY", "CREDIT", "210000", True, False, "NONE"),
        ("212000", "مصروفات مستحقة", "LIABILITY", "CREDIT", "210000", True, False, "NONE"),
        ("213000", "ضريبة قيمة مضافة مخرجات", "LIABILITY", "CREDIT", "210000", True, False, "NONE"),
        ("214000", "مستحق لشركات المجموعة", "LIABILITY", "CREDIT", "210000", True, True, "DUE_TO"),
        ("300000", "حقوق الملكية", "EQUITY", "CREDIT", None, False, False, "NONE"),
        ("311000", "رأس المال", "EQUITY", "CREDIT", "300000", True, False, "NONE"),
        ("312000", "الأرباح المحتجزة", "EQUITY", "CREDIT", "300000", True, False, "NONE"),
        ("313000", "أرباح وخسائر العام", "EQUITY", "CREDIT", "300000", True, False, "NONE"),
        ("400000", "الإيرادات", "REVENUE", "CREDIT", None, False, False, "NONE"),
        ("411000", "إيرادات الفنادق", "REVENUE", "CREDIT", "400000", True, False, "NONE"),
        ("412000", "إيرادات التطوير", "REVENUE", "CREDIT", "400000", True, False, "NONE"),
        ("413000", "إيرادات الإدارة والخدمات", "REVENUE", "CREDIT", "400000", True, False, "NONE"),
        ("419000", "إيرادات بين شركات المجموعة", "REVENUE", "CREDIT", "400000", True, True, "IC_REVENUE"),
        ("500000", "المصروفات", "EXPENSE", "DEBIT", None, False, False, "NONE"),
        ("511000", "تكلفة التشغيل", "EXPENSE", "DEBIT", "500000", True, False, "NONE"),
        ("512000", "الرواتب والأجور", "EXPENSE", "DEBIT", "500000", True, False, "NONE"),
        ("513000", "الإيجارات", "EXPENSE", "DEBIT", "500000", True, False, "NONE"),
        ("514000", "المرافق والطاقة", "EXPENSE", "DEBIT", "500000", True, False, "NONE"),
        ("515000", "مصروفات التسويق", "EXPENSE", "DEBIT", "500000", True, False, "NONE"),
        ("516000", "مصروف الإهلاك", "EXPENSE", "DEBIT", "500000", True, False, "NONE"),
        ("517000", "مصروفات بنكية", "EXPENSE", "DEBIT", "500000", True, False, "NONE"),
        ("518000", "مصروفات عمومية وإدارية", "EXPENSE", "DEBIT", "500000", True, False, "NONE"),
        ("519000", "مصروفات بين شركات المجموعة", "EXPENSE", "DEBIT", "500000", True, True, "IC_EXPENSE"),
    ]
    chart.extend([
        ("116000", "المخزون", "ASSET", "DEBIT", "110000", True, False, "NONE"),
        ("117000", "مصروفات مدفوعة مقدماً", "ASSET", "DEBIT", "110000", True, False, "NONE"),
        ("118000", "أرصدة مدينة أخرى", "ASSET", "DEBIT", "110000", True, False, "NONE"),
        ("126000", "مشروعات تحت التنفيذ", "ASSET", "DEBIT", "120000", False, False, "NONE"),
        ("126100", "أعمال إنشائية تحت التنفيذ", "ASSET", "DEBIT", "126000", True, False, "NONE"),
        ("126200", "أتعاب استشارية ورسوم مشروعات", "ASSET", "DEBIT", "126000", True, False, "NONE"),
        ("127000", "أصول غير ملموسة", "ASSET", "DEBIT", "120000", True, False, "NONE"),
        ("128000", "أصول حق استخدام", "ASSET", "DEBIT", "120000", True, False, "NONE"),
        ("129000", "استثمارات وأصول مالية طويلة الأجل", "ASSET", "DEBIT", "120000", True, False, "NONE"),
        ("215000", "ضرائب وتأمينات مستحقة", "LIABILITY", "CREDIT", "210000", True, False, "NONE"),
        ("216000", "إيرادات مؤجلة", "LIABILITY", "CREDIT", "210000", True, False, "NONE"),
        ("217000", "قروض قصيرة الأجل", "LIABILITY", "CREDIT", "210000", True, False, "NONE"),
        ("218000", "التزامات إيجار متداولة", "LIABILITY", "CREDIT", "210000", True, False, "NONE"),
        ("220000", "الالتزامات غير المتداولة", "LIABILITY", "CREDIT", "200000", False, False, "NONE"),
        ("221000", "قروض طويلة الأجل", "LIABILITY", "CREDIT", "220000", True, False, "NONE"),
        ("222000", "التزامات إيجار غير متداولة", "LIABILITY", "CREDIT", "220000", True, False, "NONE"),
        ("223000", "التزامات منافع الموظفين", "LIABILITY", "CREDIT", "220000", True, False, "NONE"),
        ("314000", "احتياطيات", "EQUITY", "CREDIT", "300000", True, False, "NONE"),
        ("315000", "الدخل الشامل الآخر", "EQUITY", "CREDIT", "300000", True, False, "NONE"),
        ("414000", "إيرادات إيجارات", "REVENUE", "CREDIT", "400000", True, False, "NONE"),
        ("415000", "إيرادات فوائد", "REVENUE", "CREDIT", "400000", True, False, "NONE"),
        ("416000", "أرباح بيع أصول", "REVENUE", "CREDIT", "400000", True, False, "NONE"),
        ("520000", "تكلفة المبيعات", "EXPENSE", "DEBIT", "500000", True, False, "NONE"),
        ("521000", "تكاليف التمويل", "EXPENSE", "DEBIT", "500000", True, False, "NONE"),
        ("522000", "مصروف ضريبة الدخل", "EXPENSE", "DEBIT", "500000", True, False, "NONE"),
        ("523000", "خسائر اضمحلال", "EXPENSE", "DEBIT", "500000", True, False, "NONE"),
        ("524000", "خسائر بيع أصول", "EXPENSE", "DEBIT", "500000", True, False, "NONE"),
    ])
    names_en = {
        "100000":"Assets","110000":"Current Assets","111000":"Cash on Hand","112000":"Banks","113000":"Trade Receivables","114000":"Input VAT","115000":"Due from Group Companies","116000":"Inventories","117000":"Prepayments","118000":"Other Receivables",
        "120000":"Non-current Assets","121000":"Land and Buildings","122000":"Furniture and Equipment","123000":"Motor Vehicles","124000":"Accumulated Depreciation","125000":"Investments in Subsidiaries","126000":"Construction in Progress","126100":"Construction Work in Progress","126200":"Project Consultancy and Fees","127000":"Intangible Assets","128000":"Right-of-use Assets","129000":"Long-term Investments and Financial Assets",
        "200000":"Liabilities","210000":"Current Liabilities","211000":"Trade Payables","212000":"Accrued Expenses","213000":"Output VAT","214000":"Due to Group Companies","215000":"Taxes and Social Insurance Payable","216000":"Deferred Revenue","217000":"Short-term Loans","218000":"Current Lease Liabilities","220000":"Non-current Liabilities","221000":"Long-term Loans","222000":"Non-current Lease Liabilities","223000":"Employee Benefit Obligations",
        "300000":"Equity","311000":"Share Capital","312000":"Retained Earnings","313000":"Current Year Profit or Loss","314000":"Reserves","315000":"Other Comprehensive Income",
        "400000":"Revenue","411000":"Hotel Revenue","412000":"Development Revenue","413000":"Management and Service Revenue","414000":"Rental Revenue","415000":"Interest Income","416000":"Gain on Disposal of Assets","419000":"Intercompany Revenue",
        "500000":"Expenses","511000":"Operating Costs","512000":"Salaries and Wages","513000":"Rent Expense","514000":"Utilities and Energy","515000":"Marketing Expenses","516000":"Depreciation Expense","517000":"Bank Charges","518000":"General and Administrative Expenses","519000":"Intercompany Expenses","520000":"Cost of Sales","521000":"Finance Costs","522000":"Income Tax Expense","523000":"Impairment Losses","524000":"Loss on Disposal of Assets",
    }
    ids: dict[str, UUID] = {}
    for code, name, cls, normal, _parent, postable, inter, role in chart:
        row = conn.execute(
            """
            INSERT INTO erp.group_accounts
                (group_id, account_code, account_name, account_name_en, account_class, normal_balance,
                 is_postable, is_intercompany, intercompany_role)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
            ON CONFLICT (group_id, account_code) DO UPDATE
              SET account_name=EXCLUDED.account_name,
                  account_name_en=EXCLUDED.account_name_en,
                  account_class=EXCLUDED.account_class,
                  normal_balance=EXCLUDED.normal_balance,
                  is_postable=EXCLUDED.is_postable,
                  is_intercompany=EXCLUDED.is_intercompany,
                  intercompany_role=EXCLUDED.intercompany_role,
                  is_active=TRUE
            RETURNING group_account_id
            """,
            (group_id, code, name, names_en[code], cls, normal, postable, inter, role),
        ).fetchone()
        ids[code] = row["group_account_id"]
    for code, _name, _cls, _normal, parent, _postable, _inter, _role in chart:
        if parent:
            conn.execute(
                """UPDATE erp.group_accounts SET parent_group_account_id=%s
                   WHERE group_id=%s AND account_code=%s""",
                (ids[parent], group_id, code),
            )


def seed_cairo_group(conn: Connection, group_id: UUID) -> None:
    holding = conn.execute(
        """
        INSERT INTO erp.companies
            (group_id, company_code, company_name, legal_name, company_kind,
             ownership_percent, functional_currency)
        VALUES (%s,'HOLD','Cairo Group Holding','Cairo Group Holding','HOLDING',100,'EGP')
        ON CONFLICT (group_id, company_code) DO UPDATE
          SET company_name=EXCLUDED.company_name, legal_name=EXCLUDED.legal_name, is_active=TRUE
        RETURNING company_id
        """,
        (group_id,),
    ).fetchone()
    company_specs = [
        ("HOTE", "شركة الفنادق", "SUBSIDIARY"),
        ("DEV", "شركة التطوير", "SUBSIDIARY"),
        ("MGT", "شركة الإدارة", "SUBSIDIARY"),
        ("ELIM", "استبعادات المجموعة", "ELIMINATION"),
    ]
    for code, name, kind in company_specs:
        conn.execute(
            """
            INSERT INTO erp.companies
                (group_id, company_code, company_name, legal_name, company_kind,
                 parent_company_id, ownership_percent, functional_currency)
            VALUES (%s,%s,%s,%s,%s,%s,100,'EGP')
            ON CONFLICT (group_id, company_code) DO UPDATE
              SET company_name=EXCLUDED.company_name, legal_name=EXCLUDED.legal_name,
                  parent_company_id=EXCLUDED.parent_company_id, is_active=TRUE
            """,
            (group_id, code, name, name, kind, holding["company_id"]),
        )

    companies = conn.execute(
        "SELECT company_id, company_code FROM erp.companies WHERE group_id=%s AND is_active=TRUE",
        (group_id,),
    ).fetchall()
    postable = conn.execute(
        """
        SELECT group_account_id, account_code, account_name, account_name_en
        FROM erp.group_accounts WHERE group_id=%s AND is_postable=TRUE AND is_active=TRUE
        """,
        (group_id,),
    ).fetchall()
    current_year = date.today().year
    for company in companies:
        for account in postable:
            conn.execute(
                """
                INSERT INTO erp.accounts
                    (group_id, company_id, group_account_id, local_account_code, local_account_name, local_account_name_en)
                VALUES (%s,%s,%s,%s,%s,%s)
                ON CONFLICT (company_id, local_account_code) DO UPDATE
                  SET group_account_id=EXCLUDED.group_account_id,
                      local_account_name_en=CASE
                        WHEN accounts.local_account_name_en IS NULL
                          OR accounts.local_account_name_en=accounts.local_account_name
                        THEN EXCLUDED.local_account_name_en
                        ELSE accounts.local_account_name_en
                      END,
                      is_active=TRUE
                """,
                (group_id, company["company_id"], account["group_account_id"], account["account_code"], account["account_name"], account["account_name_en"]),
            )
        conn.execute(
            """
            INSERT INTO erp.branches (group_id, company_id, branch_code, branch_name)
            VALUES (%s,%s,'MAIN','المركز الرئيسي')
            ON CONFLICT (company_id, branch_code) DO NOTHING
            """,
            (group_id, company["company_id"]),
        )
        conn.execute(
            """
            INSERT INTO erp.cost_centers (group_id, company_id, center_code, center_name)
            VALUES (%s,%s,'GENERAL','مركز تكلفة عام')
            ON CONFLICT (company_id, center_code) DO NOTHING
            """,
            (group_id, company["company_id"]),
        )
        fy = conn.execute(
            """
            INSERT INTO erp.fiscal_years
                (group_id, company_id, year_name, start_date, end_date)
            VALUES (%s,%s,%s,%s,%s)
            ON CONFLICT (company_id, year_name) DO UPDATE SET year_name=EXCLUDED.year_name
            RETURNING fiscal_year_id
            """,
            (group_id, company["company_id"], str(current_year), date(current_year, 1, 1), date(current_year, 12, 31)),
        ).fetchone()
        conn.execute("SELECT erp.create_monthly_periods(%s)", (fy["fiscal_year_id"],))


def migrate_and_seed() -> None:
    database_dir = Path(__file__).resolve().parent.parent / "database"
    with pool.connection() as conn:
        if AUTO_MIGRATE:
            for schema_path in sorted(database_dir.glob("*.sql")):
                conn.execute(schema_path.read_text(encoding="utf-8"))
                conn.commit()
            conn.execute("ALTER TABLE erp.group_accounts ADD COLUMN IF NOT EXISTS account_name_en VARCHAR(250)")
            conn.execute("ALTER TABLE erp.accounts ADD COLUMN IF NOT EXISTS local_account_name_en VARCHAR(250)")
            conn.execute("UPDATE erp.group_accounts SET account_name_en=account_name WHERE account_name_en IS NULL")
            conn.execute("UPDATE erp.accounts SET local_account_name_en=local_account_name WHERE local_account_name_en IS NULL")
            conn.commit()

        admin_email = os.environ.get("ADMIN_EMAIL", "admin@example.com").lower().strip()
        admin_password = os.environ.get("ADMIN_PASSWORD", "")
        if len(admin_password) < 12:
            raise RuntimeError("ADMIN_PASSWORD must be at least 12 characters")
        group_code = os.environ.get("ADMIN_GROUP_CODE", "GROUP001").strip()
        group_name = os.environ.get("ADMIN_GROUP_NAME", "Cairo Group Holding").strip()

        group = conn.execute(
            """
            INSERT INTO erp.corporate_groups
                (group_code, group_name, presentation_currency, country_code, fiscal_year_start_month)
            VALUES (%s,%s,'EGP','EG',1)
            ON CONFLICT (group_code) DO UPDATE
              SET group_name=EXCLUDED.group_name, presentation_currency='EGP', country_code='EG', fiscal_year_start_month=1
            RETURNING group_id
            """,
            (group_code, group_name),
        ).fetchone()
        seed_default_chart(conn, group["group_id"])
        seed_cairo_group(conn, group["group_id"])

        existing = conn.execute(
            "SELECT user_id FROM erp.app_users WHERE group_id=%s AND LOWER(email)=LOWER(%s)",
            (group["group_id"], admin_email),
        ).fetchone()
        if not existing:
            conn.execute(
                """
                INSERT INTO erp.app_users
                    (group_id, email, password_hash, is_group_admin, full_name, role_code)
                VALUES (%s,%s,%s,TRUE,'مدير المجموعة','GROUP_ADMIN')
                """,
                (group["group_id"], admin_email, hash_password(admin_password)),
            )
        conn.commit()


@asynccontextmanager
async def lifespan(_: FastAPI):
    pool.open(wait=True)
    migrate_and_seed()
    yield
    pool.close()


app = FastAPI(
    title="Cairo Group Holding ERP",
    version="0.4.0",
    description="Multi-company cloud accounting: GL, AR, AP, banks and fixed assets",
    lifespan=lifespan,
)


@app.middleware("http")
async def enforce_role_permissions(request: Request, call_next):
    """Enforce the assigned operational role before a write reaches an endpoint."""
    if request.method in {"POST", "PATCH", "PUT", "DELETE"} and request.url.path.startswith("/api/") and request.url.path != "/api/auth/login":
        authorization = request.headers.get("authorization", "")
        if authorization.startswith("Bearer "):
            try:
                payload = jwt.decode(authorization.removeprefix("Bearer ").strip(), JWT_SECRET, algorithms=[ALGORITHM])
                if not payload.get("is_group_admin"):
                    role = payload.get("role_code", "VIEWER")
                    if role in {"REVIEWER", "VIEWER"}:
                        return JSONResponse(
                            status_code=403,
                            content={"detail": "هذه الصلاحية للعرض والمراجعة فقط / This role has read-only access"},
                        )
                    if role == "ACCOUNTANT" and request.method == "DELETE":
                        return JSONResponse(
                            status_code=403,
                            content={"detail": "الحذف يتطلب صلاحية مدير مالي / Deletion requires Finance Manager access"},
                        )
            except jwt.PyJWTError:
                pass
    return await call_next(request)


class LoginRequest(BaseModel):
    email: EmailStr
    password: str


class CompanyCreate(BaseModel):
    company_code: str = Field(min_length=1, max_length=30)
    company_name: str = Field(min_length=1, max_length=250)
    company_kind: Literal["HOLDING", "SUBSIDIARY", "ELIMINATION"]
    parent_company_id: UUID | None = None
    ownership_percent: Decimal = Field(default=Decimal("100"), gt=0, le=100)
    functional_currency: str = Field(default="EGP", min_length=3, max_length=3)


class CompanyUpdate(BaseModel):
    company_name: str = Field(min_length=1, max_length=250)


class GroupAccountCreate(BaseModel):
    account_code: str = Field(min_length=1, max_length=50)
    account_name: str = Field(min_length=1, max_length=250)
    account_name_en: str = Field(min_length=1, max_length=250)
    account_class: Literal["ASSET", "LIABILITY", "EQUITY", "REVENUE", "EXPENSE"]
    normal_balance: Literal["DEBIT", "CREDIT"]
    parent_group_account_id: UUID | None = None
    is_postable: bool = True
    is_intercompany: bool = False
    intercompany_role: str = "NONE"


class CompanyAccountCreate(BaseModel):
    company_id: UUID
    group_account_id: UUID
    local_account_code: str = Field(min_length=1, max_length=50)
    local_account_name: str = Field(min_length=1, max_length=250)
    local_account_name_en: str = Field(min_length=1, max_length=250)


class VoucherEntryCreate(BaseModel):
    account_id: UUID
    description: str | None = None
    debit_amount: Decimal = Field(default=0, ge=0)
    credit_amount: Decimal = Field(default=0, ge=0)
    counterparty_company_id: UUID | None = None
    intercompany_reference: str | None = None
    branch_id: UUID | None = None
    cost_center_id: UUID | None = None


class VoucherCreate(BaseModel):
    company_id: UUID
    voucher_no: str = Field(min_length=1, max_length=50)
    document_date: date
    posting_date: date
    description: str | None = None
    entries: list[VoucherEntryCreate] = Field(min_length=2)
    post_immediately: bool = True


OPENING_BALANCE_HEADERS = [
    "كود الحساب",
    "اسم الحساب",
    "مدين",
    "دائن",
    "كود الشركة المقابلة",
    "مرجع Intercompany",
    "كود مركز التكلفة",
    "ملاحظات",
]


def excel_decimal(value: Any) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    try:
        return money(value)
    except Exception as exc:
        raise ValueError("القيمة المالية غير صحيحة") from exc


class FiscalYearCreate(BaseModel):
    company_id: UUID
    year_name: str
    start_date: date
    end_date: date


class CostCenterCreate(BaseModel):
    company_id: UUID
    center_code: str = Field(min_length=1, max_length=30)
    center_name: str = Field(min_length=1, max_length=250)
    parent_cost_center_id: UUID | None = None


class PartyCreate(BaseModel):
    company_id: UUID
    party_code: str = Field(min_length=1, max_length=30)
    party_name: str = Field(min_length=1, max_length=250)
    party_type: Literal["CUSTOMER", "VENDOR", "BOTH"]
    tax_registration_no: str | None = None
    email: EmailStr | None = None
    phone: str | None = None
    address: str | None = None
    receivable_account_id: UUID | None = None
    payable_account_id: UUID | None = None
    credit_limit: Decimal = Field(default=0, ge=0)
    payment_terms_days: int = Field(default=0, ge=0)


class BankAccountCreate(BaseModel):
    company_id: UUID
    bank_code: str = Field(min_length=1, max_length=30)
    bank_name: str = Field(min_length=1, max_length=250)
    account_name: str = Field(min_length=1, max_length=250)
    account_number: str | None = None
    iban: str | None = None
    currency: str = Field(default="EGP", min_length=3, max_length=3)
    gl_account_id: UUID
    opening_balance: Decimal = Decimal("0")


class InvoiceLineCreate(BaseModel):
    description: str = Field(min_length=1, max_length=500)
    account_id: UUID
    quantity: Decimal = Field(default=1, gt=0)
    unit_price: Decimal = Field(default=0, ge=0)
    tax_rate: Decimal = Field(default=0, ge=0)
    cost_center_id: UUID | None = None


class InvoiceCreate(BaseModel):
    company_id: UUID
    invoice_type: Literal["SALES", "PURCHASE"]
    invoice_no: str = Field(min_length=1, max_length=50)
    party_id: UUID
    invoice_date: date
    due_date: date
    currency: str = Field(default="EGP", min_length=3, max_length=3)
    exchange_rate: Decimal = Field(default=1, gt=0)
    description: str | None = None
    control_account_id: UUID | None = None
    tax_account_id: UUID | None = None
    lines: list[InvoiceLineCreate] = Field(min_length=1)
    post_immediately: bool = True


class CashTransactionCreate(BaseModel):
    company_id: UUID
    transaction_type: Literal["RECEIPT", "PAYMENT"]
    transaction_no: str = Field(min_length=1, max_length=50)
    transaction_date: date
    bank_account_id: UUID
    party_id: UUID | None = None
    offset_account_id: UUID | None = None
    amount: Decimal = Field(gt=0)
    description: str | None = None
    reference_no: str | None = None


class AssetCategoryCreate(BaseModel):
    company_id: UUID
    category_code: str = Field(min_length=1, max_length=30)
    category_name: str = Field(min_length=1, max_length=250)
    asset_account_id: UUID
    accumulated_depreciation_account_id: UUID
    depreciation_expense_account_id: UUID
    useful_life_months: int = Field(gt=0)


class AssetCreate(BaseModel):
    company_id: UUID
    asset_code: str = Field(min_length=1, max_length=30)
    asset_name: str = Field(min_length=1, max_length=250)
    asset_category_id: UUID
    acquisition_date: date
    placed_in_service_date: date
    acquisition_cost: Decimal = Field(ge=0)
    residual_value: Decimal = Field(default=0, ge=0)
    useful_life_months: int | None = Field(default=None, gt=0)
    location: str | None = None
    notes: str | None = None


class DepreciationRunRequest(BaseModel):
    company_id: UUID
    depreciation_date: date


class UserCreate(BaseModel):
    full_name: str = Field(min_length=2, max_length=250)
    email: EmailStr
    password: str = Field(min_length=12)
    role_code: Literal["FINANCE_MANAGER", "ACCOUNTANT", "REVIEWER", "VIEWER"]
    company_id: UUID | None = None


class UserUpdate(BaseModel):
    full_name: str = Field(min_length=2, max_length=250)
    role_code: Literal["FINANCE_MANAGER", "ACCOUNTANT", "REVIEWER", "VIEWER"]
    company_id: UUID
    password: str | None = Field(default=None, min_length=12)


class CompanyResetRequest(BaseModel):
    company_id: UUID
    confirmation: str


class GroupResetRequest(BaseModel):
    confirmation: str


def get_current_user(authorization: Annotated[str | None, Header()] = None) -> dict[str, Any]:
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Missing bearer token")
    try:
        payload = jwt.decode(authorization.removeprefix("Bearer ").strip(), JWT_SECRET, algorithms=[ALGORITHM])
        payload["user_id"] = UUID(payload["sub"])
        payload["group_id"] = UUID(payload["group_id"])
        payload["company_id"] = UUID(payload["company_id"]) if payload.get("company_id") else None
        return payload
    except (jwt.PyJWTError, ValueError) as exc:
        raise HTTPException(status_code=401, detail="Invalid or expired token") from exc


def require_group_admin(user: dict[str, Any]) -> None:
    if not user["is_group_admin"]:
        raise HTTPException(status_code=403, detail="صلاحية مدير المجموعة مطلوبة")


def ensure_company_access(user: dict[str, Any], company_id: UUID) -> None:
    if not user["is_group_admin"] and user["company_id"] != company_id:
        raise HTTPException(status_code=403, detail="لا توجد صلاحية على هذه الشركة")


@app.get("/", include_in_schema=False)
def home() -> FileResponse:
    return FileResponse(Path(__file__).parent / "static" / "index.html")


@app.get("/health")
def health() -> dict[str, str]:
    with pool.connection() as conn:
        conn.execute("SELECT 1")
    return {"status": "ok", "version": "0.4.0"}


@app.post("/api/auth/login")
def login(data: LoginRequest) -> dict[str, Any]:
    with pool.connection() as conn:
        user = conn.execute(
            """
            SELECT user_id, group_id, company_id, email, password_hash,
                   is_group_admin, role_code, full_name
            FROM erp.app_users
            WHERE LOWER(email)=LOWER(%s) AND is_active=TRUE
            """,
            (data.email,),
        ).fetchone()
        if not user or not verify_password(data.password, user["password_hash"]):
            raise HTTPException(status_code=401, detail="Invalid email or password")
        conn.execute("UPDATE erp.app_users SET last_login_at=NOW() WHERE user_id=%s", (user["user_id"],))
        conn.commit()
    return {"access_token": create_token(user), "token_type": "bearer"}


@app.get("/api/me")
def me(user: Annotated[dict[str, Any], Depends(get_current_user)]) -> dict[str, Any]:
    return {
        "user_id": str(user["user_id"]), "email": user["email"],
        "group_id": str(user["group_id"]),
        "company_id": str(user["company_id"]) if user["company_id"] else None,
        "is_group_admin": user["is_group_admin"], "role_code": user.get("role_code"),
    }


@app.get("/api/dashboard")
def dashboard(user: Annotated[dict[str, Any], Depends(get_current_user)], company_id: UUID | None = None) -> dict[str, Any]:
    if company_id:
        ensure_company_access(user, company_id)
    with pool.connection() as conn:
        group_counts = conn.execute(
            """
            SELECT
              (SELECT COUNT(*) FROM erp.companies WHERE group_id=%s AND is_active) AS companies,
              (SELECT COUNT(*) FROM erp.app_users WHERE group_id=%s AND is_active) AS users,
              (SELECT COUNT(*) FROM erp.group_accounts WHERE group_id=%s AND is_active) AS group_accounts
            """,
            (user["group_id"], user["group_id"], user["group_id"]),
        ).fetchone()
        company_data = {}
        if company_id:
            company_data = conn.execute(
                """
                SELECT
                  (SELECT COUNT(*) FROM erp.parties WHERE group_id=%s AND company_id=%s AND is_active) AS parties,
                  (SELECT COUNT(*) FROM erp.invoices WHERE group_id=%s AND company_id=%s AND status='POSTED') AS invoices,
                  (SELECT COALESCE(SUM(total_amount),0) FROM erp.invoices WHERE group_id=%s AND company_id=%s AND invoice_type='SALES' AND status IN ('POSTED','PAID')) AS sales,
                  (SELECT COALESCE(SUM(total_amount),0) FROM erp.invoices WHERE group_id=%s AND company_id=%s AND invoice_type='PURCHASE' AND status IN ('POSTED','PAID')) AS purchases,
                  (SELECT COUNT(*) FROM erp.fixed_assets WHERE group_id=%s AND company_id=%s AND status='ACTIVE') AS assets,
                  (SELECT COALESCE(SUM(opening_balance),0) FROM erp.bank_accounts WHERE group_id=%s AND company_id=%s AND is_active) AS bank_opening
                """,
                (user["group_id"], company_id, user["group_id"], company_id,
                 user["group_id"], company_id, user["group_id"], company_id,
                 user["group_id"], company_id, user["group_id"], company_id),
            ).fetchone()
    return {"group": group_counts, "company": company_data}


@app.get("/api/companies")
def list_companies(user: Annotated[dict[str, Any], Depends(get_current_user)]) -> list[dict[str, Any]]:
    with pool.connection() as conn:
        if user["is_group_admin"]:
            return conn.execute(
                """SELECT company_id, company_code, company_name, company_kind,
                          parent_company_id, ownership_percent, functional_currency
                   FROM erp.companies WHERE group_id=%s AND is_active=TRUE ORDER BY company_code""",
                (user["group_id"],),
            ).fetchall()
        return conn.execute(
            """SELECT company_id, company_code, company_name, company_kind,
                      parent_company_id, ownership_percent, functional_currency
               FROM erp.companies WHERE group_id=%s AND company_id=%s AND is_active=TRUE""",
            (user["group_id"], user["company_id"]),
        ).fetchall()


@app.post("/api/companies", status_code=201)
def create_company(data: CompanyCreate, user: Annotated[dict[str, Any], Depends(get_current_user)]) -> dict[str, Any]:
    require_group_admin(user)
    if data.company_kind != "HOLDING" and data.parent_company_id is None:
        raise HTTPException(status_code=422, detail="الشركة التابعة تحتاج شركة أم")
    with pool.connection() as conn:
        row = conn.execute(
            """
            INSERT INTO erp.companies
                (group_id, company_code, company_name, legal_name, company_kind,
                 parent_company_id, ownership_percent, functional_currency)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s) RETURNING *
            """,
            (user["group_id"], data.company_code, data.company_name, data.company_name,
             data.company_kind, data.parent_company_id, data.ownership_percent,
             data.functional_currency.upper()),
        ).fetchone()
        audit(conn, user, "CREATE", "COMPANY", row["company_id"], row["company_id"])
        conn.commit()
    return row


@app.patch("/api/companies/{company_id}")
def update_company(
    company_id: UUID,
    data: CompanyUpdate,
    user: Annotated[dict[str, Any], Depends(get_current_user)],
) -> dict[str, Any]:
    require_group_admin(user)
    with pool.connection() as conn:
        row = conn.execute(
            """UPDATE erp.companies
               SET company_name=%s, legal_name=%s
               WHERE group_id=%s AND company_id=%s AND is_active=TRUE
               RETURNING company_id, company_code, company_name, company_kind,
                         parent_company_id, ownership_percent, functional_currency""",
            (data.company_name.strip(), data.company_name.strip(), user["group_id"], company_id),
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="الشركة غير موجودة")
        audit(conn, user, "UPDATE", "COMPANY", company_id, company_id, {"company_name": data.company_name.strip()})
        conn.commit()
    return row


@app.delete("/api/companies/{company_id}")
def delete_company(
    company_id: UUID,
    user: Annotated[dict[str, Any], Depends(get_current_user)],
) -> dict[str, Any]:
    require_group_admin(user)
    with pool.connection() as conn:
        company = conn.execute(
            """SELECT company_kind, company_name
               FROM erp.companies
               WHERE group_id=%s AND company_id=%s AND is_active=TRUE""",
            (user["group_id"], company_id),
        ).fetchone()
        if not company:
            raise HTTPException(status_code=404, detail="الشركة غير موجودة")
        if company["company_kind"] == "HOLDING":
            raise HTTPException(status_code=422, detail="لا يمكن حذف الشركة القابضة الرئيسية")
        child = conn.execute(
            """SELECT 1 FROM erp.companies
               WHERE group_id=%s AND parent_company_id=%s AND is_active=TRUE LIMIT 1""",
            (user["group_id"], company_id),
        ).fetchone()
        if child:
            raise HTTPException(status_code=422, detail="لا يمكن حذف شركة لها شركات تابعة نشطة")
        conn.execute(
            """UPDATE erp.companies SET is_active=FALSE
               WHERE group_id=%s AND company_id=%s""",
            (user["group_id"], company_id),
        )
        audit(conn, user, "DEACTIVATE", "COMPANY", company_id, company_id, {"company_name": company["company_name"]})
        conn.commit()
    return {"status": "deleted", "company_id": str(company_id)}


@app.post("/api/admin/reset-company")
def reset_company_data(data: CompanyResetRequest, user: Annotated[dict[str, Any], Depends(get_current_user)]) -> dict[str, Any]:
    require_group_admin(user)
    with pool.connection() as conn:
        company = conn.execute(
            "SELECT company_code, company_name FROM erp.companies WHERE group_id=%s AND company_id=%s AND is_active=TRUE",
            (user["group_id"], data.company_id),
        ).fetchone()
        if not company:
            raise HTTPException(status_code=404, detail="الشركة غير موجودة / Company not found")
        expected = f"RESET {company['company_code']}"
        if data.confirmation.strip() != expected:
            raise HTTPException(status_code=422, detail=f"اكتب {expected} للتأكيد / Type {expected} to confirm")
        try:
            counts: dict[str, int] = {}
            commands = [
                ("asset_depreciation_entries", "DELETE FROM erp.asset_depreciation_entries WHERE group_id=%s AND company_id=%s"),
                ("fixed_assets", "DELETE FROM erp.fixed_assets WHERE group_id=%s AND company_id=%s"),
                ("asset_categories", "DELETE FROM erp.asset_categories WHERE group_id=%s AND company_id=%s"),
                ("invoice_lines", """DELETE FROM erp.invoice_lines WHERE invoice_id IN
                    (SELECT invoice_id FROM erp.invoices WHERE group_id=%s AND company_id=%s)"""),
                ("invoices", "DELETE FROM erp.invoices WHERE group_id=%s AND company_id=%s"),
                ("cash_transactions", "DELETE FROM erp.cash_transactions WHERE group_id=%s AND company_id=%s"),
                ("bank_accounts", "DELETE FROM erp.bank_accounts WHERE group_id=%s AND company_id=%s"),
                ("parties", "DELETE FROM erp.parties WHERE group_id=%s AND company_id=%s"),
                ("journal_entries", "DELETE FROM erp.journal_entries WHERE group_id=%s AND company_id=%s"),
                ("journal_vouchers", "DELETE FROM erp.journal_vouchers WHERE group_id=%s AND company_id=%s"),
            ]
            for label, sql in commands:
                result = conn.execute(sql, (user["group_id"], data.company_id))
                counts[label] = result.rowcount
            audit(conn, user, "RESET", "COMPANY_DATA", data.company_id, data.company_id, counts)
            conn.commit()
        except Exception as exc:
            conn.rollback()
            raise HTTPException(status_code=400, detail=f"تعذر تصفير الشركة / Reset failed: {exc}") from exc
    return {"status": "reset", "company_id": str(data.company_id), "deleted": counts}


@app.post("/api/admin/reset-all")
def reset_all_company_data(data: GroupResetRequest, user: Annotated[dict[str, Any], Depends(get_current_user)]) -> dict[str, Any]:
    require_group_admin(user)
    if data.confirmation.strip() != "RESET ALL":
        raise HTTPException(status_code=422, detail="اكتب RESET ALL للتأكيد / Type RESET ALL to confirm")
    with pool.connection() as conn:
        try:
            counts: dict[str, int] = {}
            commands = [
                ("asset_depreciation_entries", "DELETE FROM erp.asset_depreciation_entries WHERE group_id=%s"),
                ("fixed_assets", "DELETE FROM erp.fixed_assets WHERE group_id=%s"),
                ("asset_categories", "DELETE FROM erp.asset_categories WHERE group_id=%s"),
                ("invoice_lines", "DELETE FROM erp.invoice_lines WHERE invoice_id IN (SELECT invoice_id FROM erp.invoices WHERE group_id=%s)"),
                ("invoices", "DELETE FROM erp.invoices WHERE group_id=%s"),
                ("cash_transactions", "DELETE FROM erp.cash_transactions WHERE group_id=%s"),
                ("bank_accounts", "DELETE FROM erp.bank_accounts WHERE group_id=%s"),
                ("parties", "DELETE FROM erp.parties WHERE group_id=%s"),
                ("journal_entries", "DELETE FROM erp.journal_entries WHERE group_id=%s"),
                ("journal_vouchers", "DELETE FROM erp.journal_vouchers WHERE group_id=%s"),
            ]
            for label, sql in commands:
                result = conn.execute(sql, (user["group_id"],))
                counts[label] = result.rowcount
            audit(conn, user, "RESET", "ALL_COMPANY_DATA", details=counts)
            conn.commit()
        except Exception as exc:
            conn.rollback()
            raise HTTPException(status_code=400, detail=f"تعذر تصفير البرنامج / Reset failed: {exc}") from exc
    return {"status": "reset_all", "deleted": counts}


@app.get("/api/group-accounts")
def list_group_accounts(user: Annotated[dict[str, Any], Depends(get_current_user)]) -> list[dict[str, Any]]:
    with pool.connection() as conn:
        return conn.execute(
            """SELECT group_account_id, account_code, account_name, account_name_en, account_class,
                      normal_balance, parent_group_account_id, is_postable,
                      is_intercompany, intercompany_role
               FROM erp.group_accounts WHERE group_id=%s AND is_active=TRUE ORDER BY account_code""",
            (user["group_id"],),
        ).fetchall()


@app.post("/api/group-accounts", status_code=201)
def create_group_account(data: GroupAccountCreate, user: Annotated[dict[str, Any], Depends(get_current_user)]) -> dict[str, Any]:
    require_group_admin(user)
    if data.is_intercompany and data.intercompany_role == "NONE":
        raise HTTPException(status_code=422, detail="حدد نوع حساب Intercompany")
    with pool.connection() as conn:
        row = conn.execute(
            """
            INSERT INTO erp.group_accounts
                (group_id, account_code, account_name, account_name_en, account_class, normal_balance,
                 parent_group_account_id, is_postable, is_intercompany, intercompany_role)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING *
            """,
            (user["group_id"], data.account_code, data.account_name, data.account_name_en, data.account_class,
             data.normal_balance, data.parent_group_account_id, data.is_postable,
             data.is_intercompany, data.intercompany_role),
        ).fetchone()
        audit(conn, user, "CREATE", "GROUP_ACCOUNT", row["group_account_id"])
        conn.commit()
    return row


@app.get("/api/accounts")
def list_accounts(user: Annotated[dict[str, Any], Depends(get_current_user)], company_id: UUID) -> list[dict[str, Any]]:
    ensure_company_access(user, company_id)
    with pool.connection() as conn:
        return conn.execute(
            """
            SELECT a.account_id, a.local_account_code, a.local_account_name, a.local_account_name_en,
                   ga.account_code AS group_account_code, ga.account_name AS group_account_name,
                   ga.account_name_en AS group_account_name_en,
                   ga.account_class, ga.normal_balance, ga.is_intercompany, ga.intercompany_role
            FROM erp.accounts a
            JOIN erp.group_accounts ga ON ga.group_account_id=a.group_account_id AND ga.group_id=a.group_id
            WHERE a.group_id=%s AND a.company_id=%s AND a.is_active=TRUE
            ORDER BY a.local_account_code
            """,
            (user["group_id"], company_id),
        ).fetchall()


@app.post("/api/accounts", status_code=201)
def create_account(data: CompanyAccountCreate, user: Annotated[dict[str, Any], Depends(get_current_user)]) -> dict[str, Any]:
    ensure_company_access(user, data.company_id)
    with pool.connection() as conn:
        row = conn.execute(
            """
            INSERT INTO erp.accounts
                (group_id, company_id, group_account_id, local_account_code, local_account_name, local_account_name_en)
            VALUES (%s,%s,%s,%s,%s,%s) RETURNING *
            """,
            (user["group_id"], data.company_id, data.group_account_id,
             data.local_account_code, data.local_account_name, data.local_account_name_en),
        ).fetchone()
        audit(conn, user, "CREATE", "ACCOUNT", row["account_id"], data.company_id)
        conn.commit()
    return row


@app.get("/api/branches")
def list_branches(user: Annotated[dict[str, Any], Depends(get_current_user)], company_id: UUID) -> list[dict[str, Any]]:
    ensure_company_access(user, company_id)
    with pool.connection() as conn:
        return conn.execute(
            "SELECT branch_id, branch_code, branch_name, address FROM erp.branches WHERE group_id=%s AND company_id=%s AND is_active ORDER BY branch_code",
            (user["group_id"], company_id),
        ).fetchall()


@app.get("/api/cost-centers")
def list_cost_centers(user: Annotated[dict[str, Any], Depends(get_current_user)], company_id: UUID) -> list[dict[str, Any]]:
    ensure_company_access(user, company_id)
    with pool.connection() as conn:
        return conn.execute(
            """SELECT cost_center_id, center_code, center_name, parent_cost_center_id
               FROM erp.cost_centers WHERE group_id=%s AND company_id=%s AND is_active ORDER BY center_code""",
            (user["group_id"], company_id),
        ).fetchall()


@app.post("/api/cost-centers", status_code=201)
def create_cost_center(data: CostCenterCreate, user: Annotated[dict[str, Any], Depends(get_current_user)]) -> dict[str, Any]:
    ensure_company_access(user, data.company_id)
    with pool.connection() as conn:
        row = conn.execute(
            """INSERT INTO erp.cost_centers
                   (group_id, company_id, center_code, center_name, parent_cost_center_id)
               VALUES (%s,%s,%s,%s,%s) RETURNING *""",
            (user["group_id"], data.company_id, data.center_code, data.center_name, data.parent_cost_center_id),
        ).fetchone()
        audit(conn, user, "CREATE", "COST_CENTER", row["cost_center_id"], data.company_id)
        conn.commit()
    return row


@app.get("/api/fiscal-years")
def list_fiscal_years(user: Annotated[dict[str, Any], Depends(get_current_user)], company_id: UUID) -> list[dict[str, Any]]:
    ensure_company_access(user, company_id)
    with pool.connection() as conn:
        return conn.execute(
            """SELECT fy.fiscal_year_id, fy.year_name, fy.start_date, fy.end_date, fy.status,
                      COUNT(fp.period_id) AS periods,
                      COUNT(*) FILTER (WHERE fp.status='OPEN') AS open_periods
               FROM erp.fiscal_years fy LEFT JOIN erp.fiscal_periods fp ON fp.fiscal_year_id=fy.fiscal_year_id
               WHERE fy.group_id=%s AND fy.company_id=%s
               GROUP BY fy.fiscal_year_id ORDER BY fy.start_date DESC""",
            (user["group_id"], company_id),
        ).fetchall()


@app.post("/api/fiscal-years", status_code=201)
def create_fiscal_year(data: FiscalYearCreate, user: Annotated[dict[str, Any], Depends(get_current_user)]) -> dict[str, Any]:
    ensure_company_access(user, data.company_id)
    with pool.connection() as conn:
        row = conn.execute(
            """INSERT INTO erp.fiscal_years
                   (group_id, company_id, year_name, start_date, end_date)
               VALUES (%s,%s,%s,%s,%s) RETURNING *""",
            (user["group_id"], data.company_id, data.year_name, data.start_date, data.end_date),
        ).fetchone()
        conn.execute("SELECT erp.create_monthly_periods(%s)", (row["fiscal_year_id"],))
        audit(conn, user, "CREATE", "FISCAL_YEAR", row["fiscal_year_id"], data.company_id)
        conn.commit()
    return row


@app.get("/api/fiscal-periods")
def list_fiscal_periods(user: Annotated[dict[str, Any], Depends(get_current_user)], company_id: UUID, fiscal_year_id: UUID | None = None) -> list[dict[str, Any]]:
    ensure_company_access(user, company_id)
    with pool.connection() as conn:
        return conn.execute(
            """SELECT period_id, fiscal_year_id, period_no, period_name, start_date, end_date, status
               FROM erp.fiscal_periods
               WHERE group_id=%s AND company_id=%s AND (%s IS NULL OR fiscal_year_id=%s)
               ORDER BY start_date""",
            (user["group_id"], company_id, fiscal_year_id, fiscal_year_id),
        ).fetchall()


@app.patch("/api/fiscal-periods/{period_id}")
def update_period(period_id: UUID, new_status: Literal["OPEN", "CLOSED", "LOCKED"], user: Annotated[dict[str, Any], Depends(get_current_user)]) -> dict[str, Any]:
    with pool.connection() as conn:
        period = conn.execute("SELECT company_id FROM erp.fiscal_periods WHERE period_id=%s AND group_id=%s", (period_id, user["group_id"])).fetchone()
        if not period:
            raise HTTPException(status_code=404, detail="الفترة غير موجودة")
        ensure_company_access(user, period["company_id"])
        row = conn.execute("UPDATE erp.fiscal_periods SET status=%s WHERE period_id=%s RETURNING *", (new_status, period_id)).fetchone()
        audit(conn, user, "UPDATE_STATUS", "FISCAL_PERIOD", period_id, period["company_id"], {"status": new_status})
        conn.commit()
    return row


@app.get("/api/vouchers")
def list_vouchers(user: Annotated[dict[str, Any], Depends(get_current_user)], company_id: UUID, limit: int = Query(100, ge=1, le=500)) -> list[dict[str, Any]]:
    ensure_company_access(user, company_id)
    with pool.connection() as conn:
        return conn.execute(
            """SELECT voucher_id, voucher_no, voucher_type, status, document_date, posting_date,
                      description, source_module, external_reference, created_at, posted_at,
                      (SELECT COALESCE(SUM(debit_amount),0) FROM erp.journal_entries e WHERE e.voucher_id=v.voucher_id) AS amount
               FROM erp.journal_vouchers v
               WHERE group_id=%s AND company_id=%s ORDER BY posting_date DESC, created_at DESC LIMIT %s""",
            (user["group_id"], company_id, limit),
        ).fetchall()


@app.post("/api/vouchers", status_code=201)
def create_voucher(data: VoucherCreate, user: Annotated[dict[str, Any], Depends(get_current_user)]) -> dict[str, Any]:
    ensure_company_access(user, data.company_id)
    with pool.connection() as conn:
        try:
            voucher = create_voucher_db(
                conn, user=user, company_id=data.company_id, voucher_no=data.voucher_no,
                document_date=data.document_date, posting_date=data.posting_date,
                description=data.description, entries=[e.model_dump() for e in data.entries],
                post_immediately=data.post_immediately,
            )
            audit(conn, user, "CREATE", "VOUCHER", voucher["voucher_id"], data.company_id, {"status": voucher["status"]})
            conn.commit()
            return voucher
        except HTTPException:
            conn.rollback(); raise
        except Exception as exc:
            conn.rollback(); raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.get("/api/opening-balances/template")
def download_opening_balance_template(
    company_id: UUID,
    user: Annotated[dict[str, Any], Depends(get_current_user)],
) -> StreamingResponse:
    ensure_company_access(user, company_id)
    with pool.connection() as conn:
        company = conn.execute(
            """SELECT company_code, company_name, functional_currency
               FROM erp.companies
               WHERE group_id=%s AND company_id=%s AND is_active=TRUE""",
            (user["group_id"], company_id),
        ).fetchone()
        accounts = conn.execute(
            """SELECT a.local_account_code, a.local_account_name, a.local_account_name_en, ga.is_intercompany
               FROM erp.accounts a
               JOIN erp.group_accounts ga
                 ON ga.group_account_id=a.group_account_id AND ga.group_id=a.group_id
               WHERE a.group_id=%s AND a.company_id=%s AND a.is_active=TRUE
                 AND ga.is_postable=TRUE
               ORDER BY a.local_account_code""",
            (user["group_id"], company_id),
        ).fetchall()
        companies = conn.execute(
            """SELECT company_code, company_name
               FROM erp.companies
               WHERE group_id=%s AND company_id<>%s AND is_active=TRUE
               ORDER BY company_code""",
            (user["group_id"], company_id),
        ).fetchall()
        cost_centers = conn.execute(
            """SELECT center_code, center_name
               FROM erp.cost_centers
               WHERE group_id=%s AND company_id=%s AND is_active=TRUE
               ORDER BY center_code""",
            (user["group_id"], company_id),
        ).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "الأرصدة الافتتاحية"
    ws.sheet_view.rightToLeft = True
    ws.append(OPENING_BALANCE_HEADERS)
    header_fill = PatternFill("solid", fgColor="0B6B61")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    for account in accounts:
        ws.append([
            account["local_account_code"],
            f"{account['local_account_name']} / {account['local_account_name_en']}",
            None,
            None,
            None,
            None,
            None,
            "Intercompany" if account["is_intercompany"] else None,
        ])
    widths = [18, 38, 16, 16, 24, 24, 22, 34]
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + index)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{max(ws.max_row, 2)}"

    info = wb.create_sheet("تعليمات")
    info.sheet_view.rightToLeft = True
    info.append(["قالب الأرصدة الافتتاحية", company["company_name"]])
    info.append(["كود الشركة", company["company_code"]])
    info.append(["العملة الوظيفية", company["functional_currency"]])
    info.append(["تعليمات", "أدخل مبلغاً في مدين أو دائن فقط. احذف السطور غير المستخدمة أو اترك مبالغها فارغة."])
    info.append(["Intercompany", "أدخل كود الشركة المقابلة ومرجع المعاملة للحسابات المتبادلة فقط."])
    info.column_dimensions["A"].width = 24
    info.column_dimensions["B"].width = 90
    info["A1"].font = info["B1"].font = Font(bold=True, color="FFFFFF")
    info["A1"].fill = info["B1"].fill = header_fill

    refs = wb.create_sheet("القوائم المرجعية")
    refs.sheet_view.rightToLeft = True
    refs.append(["كود الشركة المقابلة", "اسم الشركة", "", "كود مركز التكلفة", "اسم مركز التكلفة"])
    for i in range(max(len(companies), len(cost_centers))):
        company_row = companies[i] if i < len(companies) else {}
        center_row = cost_centers[i] if i < len(cost_centers) else {}
        refs.append([
            company_row.get("company_code"),
            company_row.get("company_name"),
            "",
            center_row.get("center_code"),
            center_row.get("center_name"),
        ])
    refs.sheet_state = "hidden"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    safe_code = "".join(ch for ch in company["company_code"] if ch.isalnum() or ch in "-_")
    filename = f"opening-balances-{safe_code or 'company'}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.post("/api/opening-balances/preview")
async def preview_opening_balances(
    request: Request,
    company_id: UUID,
    user: Annotated[dict[str, Any], Depends(get_current_user)],
) -> dict[str, Any]:
    ensure_company_access(user, company_id)
    content = await request.body()
    if not content:
        raise HTTPException(status_code=422, detail="اختر ملف Excel أولاً")
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="حجم الملف يجب ألا يتجاوز 10 ميجابايت")
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb["الأرصدة الافتتاحية"] if "الأرصدة الافتتاحية" in wb.sheetnames else wb.active
    except Exception as exc:
        raise HTTPException(status_code=422, detail="تعذر قراءة الملف. استخدم قالب Excel الصادر من النظام.") from exc

    header_values = [str(value).strip() if value is not None else "" for value in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    header_map = {name: index for index, name in enumerate(header_values)}
    missing = [name for name in ("كود الحساب", "مدين", "دائن") if name not in header_map]
    if missing:
        raise HTTPException(status_code=422, detail=f"أعمدة مطلوبة غير موجودة: {', '.join(missing)}")

    with pool.connection() as conn:
        account_rows = conn.execute(
            """SELECT a.account_id, a.local_account_code, a.local_account_name, a.local_account_name_en,
                      ga.is_intercompany
               FROM erp.accounts a
               JOIN erp.group_accounts ga
                 ON ga.group_account_id=a.group_account_id AND ga.group_id=a.group_id
               WHERE a.group_id=%s AND a.company_id=%s AND a.is_active=TRUE
                 AND ga.is_postable=TRUE""",
            (user["group_id"], company_id),
        ).fetchall()
        company_rows = conn.execute(
            """SELECT company_id, company_code, company_name
               FROM erp.companies
               WHERE group_id=%s AND is_active=TRUE""",
            (user["group_id"],),
        ).fetchall()
        center_rows = conn.execute(
            """SELECT cost_center_id, center_code, center_name
               FROM erp.cost_centers
               WHERE group_id=%s AND company_id=%s AND is_active=TRUE""",
            (user["group_id"], company_id),
        ).fetchall()

    accounts = {str(row["local_account_code"]).strip(): row for row in account_rows}
    companies = {str(row["company_code"]).strip().upper(): row for row in company_rows}
    centers = {str(row["center_code"]).strip().upper(): row for row in center_rows}
    current_company = next((row for row in company_rows if row["company_id"] == company_id), None)

    def value(row: tuple[Any, ...], name: str) -> Any:
        index = header_map.get(name)
        return row[index] if index is not None and index < len(row) else None

    preview: list[dict[str, Any]] = []
    total_debit = Decimal("0")
    total_credit = Decimal("0")
    for excel_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        code = str(value(row, "كود الحساب") or "").strip()
        raw_debit = value(row, "مدين")
        raw_credit = value(row, "دائن")
        if not code and raw_debit in (None, "") and raw_credit in (None, ""):
            continue
        errors: list[str] = []
        account = accounts.get(code)
        if not account:
            errors.append("كود الحساب غير موجود أو غير قابل للترحيل")
        try:
            debit = excel_decimal(raw_debit)
            credit = excel_decimal(raw_credit)
        except ValueError as exc:
            debit = credit = Decimal("0")
            errors.append(str(exc))
        if debit < 0 or credit < 0:
            errors.append("لا يمكن إدخال مبلغ سالب")
        if (debit > 0) == (credit > 0):
            errors.append("أدخل مبلغاً في مدين أو دائن فقط")

        counterparty_code = str(value(row, "كود الشركة المقابلة") or "").strip().upper()
        counterparty = companies.get(counterparty_code) if counterparty_code else None
        ic_reference = str(value(row, "مرجع Intercompany") or "").strip()
        if counterparty_code and not counterparty:
            errors.append("كود الشركة المقابلة غير موجود")
        if counterparty and current_company and counterparty["company_id"] == current_company["company_id"]:
            errors.append("الشركة المقابلة لا يمكن أن تكون نفس الشركة")
        if account and account["is_intercompany"]:
            if not counterparty:
                errors.append("الشركة المقابلة مطلوبة لحساب Intercompany")
            if not ic_reference:
                errors.append("مرجع Intercompany مطلوب")
        elif counterparty_code or ic_reference:
            errors.append("بيانات Intercompany غير مسموحة لهذا الحساب")

        center_code = str(value(row, "كود مركز التكلفة") or "").strip().upper()
        center = centers.get(center_code) if center_code else None
        if center_code and not center:
            errors.append("كود مركز التكلفة غير موجود")
        if not errors:
            total_debit += debit
            total_credit += credit
        preview.append({
            "excel_row": excel_row,
            "account_id": str(account["account_id"]) if account else None,
            "account_code": code,
            "account_name": account["local_account_name"] if account else str(value(row, "اسم الحساب") or ""),
            "account_name_en": account["local_account_name_en"] if account else "",
            "debit_amount": str(debit),
            "credit_amount": str(credit),
            "counterparty_company_id": str(counterparty["company_id"]) if counterparty else None,
            "counterparty_company_code": counterparty_code or None,
            "intercompany_reference": ic_reference or None,
            "cost_center_id": str(center["cost_center_id"]) if center else None,
            "description": str(value(row, "ملاحظات") or "").strip() or None,
            "errors": errors,
            "valid": not errors,
        })
    if not preview:
        raise HTTPException(status_code=422, detail="الملف لا يحتوي على أرصدة لإظهارها")
    difference = total_debit - total_credit
    return {
        "rows": preview,
        "valid_rows": sum(1 for row in preview if row["valid"]),
        "invalid_rows": sum(1 for row in preview if not row["valid"]),
        "total_debit": str(money(total_debit)),
        "total_credit": str(money(total_credit)),
        "difference": str(money(difference)),
        "balanced": difference == 0 and total_debit > 0,
    }


@app.get("/api/parties")
def list_parties(user: Annotated[dict[str, Any], Depends(get_current_user)], company_id: UUID, party_type: str | None = None) -> list[dict[str, Any]]:
    ensure_company_access(user, company_id)
    with pool.connection() as conn:
        return conn.execute(
            """SELECT party_id, party_code, party_name, party_type, tax_registration_no,
                      email, phone, address, receivable_account_id, payable_account_id,
                      credit_limit, payment_terms_days
               FROM erp.parties
               WHERE group_id=%s AND company_id=%s AND is_active=TRUE
                 AND (%s IS NULL OR party_type=%s OR party_type='BOTH')
               ORDER BY party_code""",
            (user["group_id"], company_id, party_type, party_type),
        ).fetchall()


@app.post("/api/parties", status_code=201)
def create_party(data: PartyCreate, user: Annotated[dict[str, Any], Depends(get_current_user)]) -> dict[str, Any]:
    ensure_company_access(user, data.company_id)
    if data.party_type in ("CUSTOMER", "BOTH") and not data.receivable_account_id:
        raise HTTPException(status_code=422, detail="حساب العملاء مطلوب")
    if data.party_type in ("VENDOR", "BOTH") and not data.payable_account_id:
        raise HTTPException(status_code=422, detail="حساب الموردين مطلوب")
    with pool.connection() as conn:
        if data.receivable_account_id:
            ensure_account(conn, user["group_id"], data.company_id, data.receivable_account_id)
        if data.payable_account_id:
            ensure_account(conn, user["group_id"], data.company_id, data.payable_account_id)
        row = conn.execute(
            """INSERT INTO erp.parties
                   (group_id, company_id, party_code, party_name, party_type,
                    tax_registration_no, email, phone, address, receivable_account_id,
                    payable_account_id, credit_limit, payment_terms_days)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING *""",
            (user["group_id"], data.company_id, data.party_code, data.party_name,
             data.party_type, data.tax_registration_no, str(data.email) if data.email else None,
             data.phone, data.address, data.receivable_account_id, data.payable_account_id,
             data.credit_limit, data.payment_terms_days),
        ).fetchone()
        audit(conn, user, "CREATE", "PARTY", row["party_id"], data.company_id, {"type": data.party_type})
        conn.commit()
    return row


@app.get("/api/bank-accounts")
def list_bank_accounts(user: Annotated[dict[str, Any], Depends(get_current_user)], company_id: UUID) -> list[dict[str, Any]]:
    ensure_company_access(user, company_id)
    with pool.connection() as conn:
        return conn.execute(
            """SELECT b.bank_account_id, b.bank_code, b.bank_name, b.account_name,
                      b.account_number, b.iban, b.currency, b.opening_balance,
                      b.gl_account_id, a.local_account_code, a.local_account_name, a.local_account_name_en,
                      b.opening_balance + COALESCE(SUM(CASE WHEN ct.transaction_type='RECEIPT' THEN ct.amount ELSE -ct.amount END),0) AS current_balance
               FROM erp.bank_accounts b
               JOIN erp.accounts a ON a.account_id=b.gl_account_id
               LEFT JOIN erp.cash_transactions ct ON ct.bank_account_id=b.bank_account_id AND ct.status='POSTED'
               WHERE b.group_id=%s AND b.company_id=%s AND b.is_active
               GROUP BY b.bank_account_id, a.account_id ORDER BY b.bank_code""",
            (user["group_id"], company_id),
        ).fetchall()


@app.post("/api/bank-accounts", status_code=201)
def create_bank_account(data: BankAccountCreate, user: Annotated[dict[str, Any], Depends(get_current_user)]) -> dict[str, Any]:
    ensure_company_access(user, data.company_id)
    with pool.connection() as conn:
        ensure_account(conn, user["group_id"], data.company_id, data.gl_account_id)
        row = conn.execute(
            """INSERT INTO erp.bank_accounts
                   (group_id, company_id, bank_code, bank_name, account_name,
                    account_number, iban, currency, gl_account_id, opening_balance)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING *""",
            (user["group_id"], data.company_id, data.bank_code, data.bank_name,
             data.account_name, data.account_number, data.iban, data.currency.upper(),
             data.gl_account_id, data.opening_balance),
        ).fetchone()
        audit(conn, user, "CREATE", "BANK_ACCOUNT", row["bank_account_id"], data.company_id)
        conn.commit()
    return row


@app.get("/api/invoices")
def list_invoices(user: Annotated[dict[str, Any], Depends(get_current_user)], company_id: UUID, invoice_type: str | None = None) -> list[dict[str, Any]]:
    ensure_company_access(user, company_id)
    with pool.connection() as conn:
        return conn.execute(
            """SELECT i.invoice_id, i.invoice_type, i.invoice_no, i.invoice_date, i.due_date,
                      i.currency, i.subtotal, i.tax_amount, i.total_amount, i.status,
                      p.party_code, p.party_name, i.description
               FROM erp.invoices i JOIN erp.parties p ON p.party_id=i.party_id
               WHERE i.group_id=%s AND i.company_id=%s AND (%s IS NULL OR i.invoice_type=%s)
               ORDER BY i.invoice_date DESC, i.created_at DESC""",
            (user["group_id"], company_id, invoice_type, invoice_type),
        ).fetchall()


@app.post("/api/invoices", status_code=201)
def create_invoice(data: InvoiceCreate, user: Annotated[dict[str, Any], Depends(get_current_user)]) -> dict[str, Any]:
    ensure_company_access(user, data.company_id)
    if data.due_date < data.invoice_date:
        raise HTTPException(status_code=422, detail="تاريخ الاستحقاق لا يسبق تاريخ الفاتورة")
    with pool.connection() as conn:
        try:
            party = conn.execute(
                "SELECT * FROM erp.parties WHERE group_id=%s AND company_id=%s AND party_id=%s AND is_active",
                (user["group_id"], data.company_id, data.party_id),
            ).fetchone()
            if not party:
                raise HTTPException(status_code=422, detail="العميل أو المورد غير موجود")
            if data.invoice_type == "SALES" and party["party_type"] not in ("CUSTOMER", "BOTH"):
                raise HTTPException(status_code=422, detail="الطرف ليس عميلاً")
            if data.invoice_type == "PURCHASE" and party["party_type"] not in ("VENDOR", "BOTH"):
                raise HTTPException(status_code=422, detail="الطرف ليس مورداً")
            control_account = data.control_account_id or (
                party["receivable_account_id"] if data.invoice_type == "SALES" else party["payable_account_id"]
            )
            if not control_account:
                raise HTTPException(status_code=422, detail="حساب العميل/المورد غير محدد")
            ensure_account(conn, user["group_id"], data.company_id, control_account)
            if data.tax_account_id:
                ensure_account(conn, user["group_id"], data.company_id, data.tax_account_id)

            calculated = []
            subtotal = Decimal("0")
            tax_total = Decimal("0")
            for line in data.lines:
                ensure_account(conn, user["group_id"], data.company_id, line.account_id)
                net = money(line.quantity * line.unit_price)
                tax = money(net * line.tax_rate / Decimal("100"))
                total = money(net + tax)
                subtotal += net; tax_total += tax
                calculated.append((line, net, tax, total))
            subtotal = money(subtotal); tax_total = money(tax_total); total_amount = money(subtotal + tax_total)
            if tax_total > 0 and not data.tax_account_id:
                raise HTTPException(status_code=422, detail="اختر حساب الضريبة لأن الفاتورة تحتوي ضريبة")

            invoice = conn.execute(
                """INSERT INTO erp.invoices
                       (group_id, company_id, invoice_type, invoice_no, party_id,
                        invoice_date, due_date, currency, exchange_rate, description,
                        subtotal, tax_amount, total_amount, control_account_id,
                        tax_account_id, status, created_by)
                   VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'DRAFT',%s)
                   RETURNING *""",
                (user["group_id"], data.company_id, data.invoice_type, data.invoice_no,
                 data.party_id, data.invoice_date, data.due_date, data.currency.upper(),
                 data.exchange_rate, data.description, subtotal, tax_total, total_amount,
                 control_account, data.tax_account_id, user["user_id"]),
            ).fetchone()
            for idx, (line, net, tax, total) in enumerate(calculated, 1):
                conn.execute(
                    """INSERT INTO erp.invoice_lines
                           (invoice_id, group_id, company_id, line_no, description,
                            account_id, quantity, unit_price, tax_rate, net_amount,
                            tax_amount, total_amount, cost_center_id)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                    (invoice["invoice_id"], user["group_id"], data.company_id, idx,
                     line.description, line.account_id, line.quantity, line.unit_price,
                     line.tax_rate, net, tax, total, line.cost_center_id),
                )

            if data.post_immediately:
                entries: list[dict[str, Any]] = []
                if data.invoice_type == "SALES":
                    entries.append({"account_id": control_account, "debit_amount": total_amount, "credit_amount": 0, "description": f"فاتورة مبيعات {data.invoice_no}"})
                    for line, net, _tax, _total in calculated:
                        entries.append({"account_id": line.account_id, "debit_amount": 0, "credit_amount": net, "description": line.description, "cost_center_id": line.cost_center_id})
                    if tax_total > 0:
                        entries.append({"account_id": data.tax_account_id, "debit_amount": 0, "credit_amount": tax_total, "description": "ضريبة مخرجات"})
                else:
                    for line, net, _tax, _total in calculated:
                        entries.append({"account_id": line.account_id, "debit_amount": net, "credit_amount": 0, "description": line.description, "cost_center_id": line.cost_center_id})
                    if tax_total > 0:
                        entries.append({"account_id": data.tax_account_id, "debit_amount": tax_total, "credit_amount": 0, "description": "ضريبة مدخلات"})
                    entries.append({"account_id": control_account, "debit_amount": 0, "credit_amount": total_amount, "description": f"فاتورة مشتريات {data.invoice_no}"})
                voucher = create_voucher_db(
                    conn, user=user, company_id=data.company_id,
                    voucher_no=f"{'SI' if data.invoice_type == 'SALES' else 'PI'}-{data.invoice_no}",
                    document_date=data.invoice_date, posting_date=data.invoice_date,
                    description=data.description or f"فاتورة {data.invoice_no}", entries=entries,
                    source_module="AR" if data.invoice_type == "SALES" else "AP",
                    external_reference=data.invoice_no, post_immediately=True,
                )
                conn.execute("UPDATE erp.invoices SET status='POSTED', voucher_id=%s WHERE invoice_id=%s", (voucher["voucher_id"], invoice["invoice_id"]))
                invoice["status"] = "POSTED"; invoice["voucher_id"] = voucher["voucher_id"]
            audit(conn, user, "CREATE", "INVOICE", invoice["invoice_id"], data.company_id, {"type": data.invoice_type, "total": str(total_amount)})
            conn.commit()
            return invoice
        except HTTPException:
            conn.rollback(); raise
        except Exception as exc:
            conn.rollback(); raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.get("/api/cash-transactions")
def list_cash_transactions(user: Annotated[dict[str, Any], Depends(get_current_user)], company_id: UUID) -> list[dict[str, Any]]:
    ensure_company_access(user, company_id)
    with pool.connection() as conn:
        return conn.execute(
            """SELECT ct.cash_transaction_id, ct.transaction_type, ct.transaction_no,
                      ct.transaction_date, ct.amount, ct.description, ct.reference_no,
                      b.bank_name, b.account_name, p.party_name, ct.status
               FROM erp.cash_transactions ct
               JOIN erp.bank_accounts b ON b.bank_account_id=ct.bank_account_id
               LEFT JOIN erp.parties p ON p.party_id=ct.party_id
               WHERE ct.group_id=%s AND ct.company_id=%s
               ORDER BY ct.transaction_date DESC, ct.created_at DESC""",
            (user["group_id"], company_id),
        ).fetchall()


@app.post("/api/cash-transactions", status_code=201)
def create_cash_transaction(data: CashTransactionCreate, user: Annotated[dict[str, Any], Depends(get_current_user)]) -> dict[str, Any]:
    ensure_company_access(user, data.company_id)
    with pool.connection() as conn:
        try:
            bank = conn.execute(
                "SELECT * FROM erp.bank_accounts WHERE group_id=%s AND company_id=%s AND bank_account_id=%s AND is_active",
                (user["group_id"], data.company_id, data.bank_account_id),
            ).fetchone()
            if not bank:
                raise HTTPException(status_code=422, detail="الحساب البنكي غير موجود")
            offset = data.offset_account_id
            party = None
            if data.party_id:
                party = conn.execute(
                    "SELECT * FROM erp.parties WHERE group_id=%s AND company_id=%s AND party_id=%s",
                    (user["group_id"], data.company_id, data.party_id),
                ).fetchone()
                if not party:
                    raise HTTPException(status_code=422, detail="الطرف غير موجود")
                if not offset:
                    offset = party["receivable_account_id"] if data.transaction_type == "RECEIPT" else party["payable_account_id"]
            if not offset:
                raise HTTPException(status_code=422, detail="حدد الحساب المقابل")
            ensure_account(conn, user["group_id"], data.company_id, offset)
            description = data.description or ("سند قبض" if data.transaction_type == "RECEIPT" else "سند صرف")
            if data.transaction_type == "RECEIPT":
                entries = [
                    {"account_id": bank["gl_account_id"], "debit_amount": data.amount, "credit_amount": 0, "description": description},
                    {"account_id": offset, "debit_amount": 0, "credit_amount": data.amount, "description": description},
                ]
            else:
                entries = [
                    {"account_id": offset, "debit_amount": data.amount, "credit_amount": 0, "description": description},
                    {"account_id": bank["gl_account_id"], "debit_amount": 0, "credit_amount": data.amount, "description": description},
                ]
            voucher = create_voucher_db(
                conn, user=user, company_id=data.company_id,
                voucher_no=f"{'RV' if data.transaction_type == 'RECEIPT' else 'PV'}-{data.transaction_no}",
                document_date=data.transaction_date, posting_date=data.transaction_date,
                description=description, entries=entries, source_module="BANK",
                external_reference=data.reference_no or data.transaction_no, post_immediately=True,
            )
            row = conn.execute(
                """INSERT INTO erp.cash_transactions
                       (group_id, company_id, transaction_type, transaction_no,
                        transaction_date, bank_account_id, party_id, offset_account_id,
                        amount, description, reference_no, status, voucher_id, created_by)
                   VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'POSTED',%s,%s) RETURNING *""",
                (user["group_id"], data.company_id, data.transaction_type,
                 data.transaction_no, data.transaction_date, data.bank_account_id,
                 data.party_id, offset, data.amount, description, data.reference_no,
                 voucher["voucher_id"], user["user_id"]),
            ).fetchone()
            audit(conn, user, "CREATE", "CASH_TRANSACTION", row["cash_transaction_id"], data.company_id, {"type": data.transaction_type, "amount": str(data.amount)})
            conn.commit()
            return row
        except HTTPException:
            conn.rollback(); raise
        except Exception as exc:
            conn.rollback(); raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.get("/api/asset-categories")
def list_asset_categories(user: Annotated[dict[str, Any], Depends(get_current_user)], company_id: UUID) -> list[dict[str, Any]]:
    ensure_company_access(user, company_id)
    with pool.connection() as conn:
        return conn.execute(
            """SELECT ac.*, aa.local_account_name AS asset_account_name,
                      ad.local_account_name AS accumulated_account_name,
                      de.local_account_name AS expense_account_name
               FROM erp.asset_categories ac
               JOIN erp.accounts aa ON aa.account_id=ac.asset_account_id
               JOIN erp.accounts ad ON ad.account_id=ac.accumulated_depreciation_account_id
               JOIN erp.accounts de ON de.account_id=ac.depreciation_expense_account_id
               WHERE ac.group_id=%s AND ac.company_id=%s AND ac.is_active ORDER BY ac.category_code""",
            (user["group_id"], company_id),
        ).fetchall()


@app.post("/api/asset-categories", status_code=201)
def create_asset_category(data: AssetCategoryCreate, user: Annotated[dict[str, Any], Depends(get_current_user)]) -> dict[str, Any]:
    ensure_company_access(user, data.company_id)
    with pool.connection() as conn:
        for account_id in (data.asset_account_id, data.accumulated_depreciation_account_id, data.depreciation_expense_account_id):
            ensure_account(conn, user["group_id"], data.company_id, account_id)
        row = conn.execute(
            """INSERT INTO erp.asset_categories
                   (group_id, company_id, category_code, category_name, asset_account_id,
                    accumulated_depreciation_account_id, depreciation_expense_account_id,
                    useful_life_months)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s) RETURNING *""",
            (user["group_id"], data.company_id, data.category_code, data.category_name,
             data.asset_account_id, data.accumulated_depreciation_account_id,
             data.depreciation_expense_account_id, data.useful_life_months),
        ).fetchone()
        audit(conn, user, "CREATE", "ASSET_CATEGORY", row["asset_category_id"], data.company_id)
        conn.commit()
    return row


@app.get("/api/assets")
def list_assets(user: Annotated[dict[str, Any], Depends(get_current_user)], company_id: UUID) -> list[dict[str, Any]]:
    ensure_company_access(user, company_id)
    with pool.connection() as conn:
        return conn.execute(
            """SELECT fa.*, ac.category_code, ac.category_name,
                      (fa.acquisition_cost-fa.residual_value)/fa.useful_life_months AS monthly_depreciation,
                      fa.acquisition_cost-fa.accumulated_depreciation AS net_book_value
               FROM erp.fixed_assets fa JOIN erp.asset_categories ac ON ac.asset_category_id=fa.asset_category_id
               WHERE fa.group_id=%s AND fa.company_id=%s ORDER BY fa.asset_code""",
            (user["group_id"], company_id),
        ).fetchall()


@app.post("/api/assets", status_code=201)
def create_asset(data: AssetCreate, user: Annotated[dict[str, Any], Depends(get_current_user)]) -> dict[str, Any]:
    ensure_company_access(user, data.company_id)
    if data.residual_value > data.acquisition_cost:
        raise HTTPException(status_code=422, detail="القيمة التخريدية لا تتجاوز تكلفة الأصل")
    with pool.connection() as conn:
        category = conn.execute(
            "SELECT * FROM erp.asset_categories WHERE group_id=%s AND company_id=%s AND asset_category_id=%s",
            (user["group_id"], data.company_id, data.asset_category_id),
        ).fetchone()
        if not category:
            raise HTTPException(status_code=422, detail="فئة الأصل غير موجودة")
        row = conn.execute(
            """INSERT INTO erp.fixed_assets
                   (group_id, company_id, asset_code, asset_name, asset_category_id,
                    acquisition_date, placed_in_service_date, acquisition_cost,
                    residual_value, useful_life_months, location, notes)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING *""",
            (user["group_id"], data.company_id, data.asset_code, data.asset_name,
             data.asset_category_id, data.acquisition_date, data.placed_in_service_date,
             data.acquisition_cost, data.residual_value,
             data.useful_life_months or category["useful_life_months"], data.location, data.notes),
        ).fetchone()
        audit(conn, user, "CREATE", "FIXED_ASSET", row["asset_id"], data.company_id)
        conn.commit()
    return row


@app.get("/api/assets/depreciation-preview")
def depreciation_preview(user: Annotated[dict[str, Any], Depends(get_current_user)], company_id: UUID, depreciation_date: date) -> list[dict[str, Any]]:
    ensure_company_access(user, company_id)
    with pool.connection() as conn:
        rows = conn.execute(
            """SELECT fa.asset_id, fa.asset_code, fa.asset_name, fa.acquisition_cost,
                      fa.residual_value, fa.accumulated_depreciation, fa.useful_life_months,
                      ac.depreciation_expense_account_id, ac.accumulated_depreciation_account_id,
                      GREATEST(LEAST((fa.acquisition_cost-fa.residual_value)/fa.useful_life_months,
                           fa.acquisition_cost-fa.residual_value-fa.accumulated_depreciation),0)::NUMERIC(20,4) AS amount
               FROM erp.fixed_assets fa JOIN erp.asset_categories ac ON ac.asset_category_id=fa.asset_category_id
               WHERE fa.group_id=%s AND fa.company_id=%s AND fa.status='ACTIVE'
                 AND fa.placed_in_service_date<=%s
                 AND (fa.last_depreciation_date IS NULL OR fa.last_depreciation_date<%s)
               ORDER BY fa.asset_code""",
            (user["group_id"], company_id, depreciation_date, depreciation_date),
        ).fetchall()
    return [r for r in rows if Decimal(str(r["amount"])) > 0]


@app.post("/api/assets/run-depreciation", status_code=201)
def run_depreciation(data: DepreciationRunRequest, user: Annotated[dict[str, Any], Depends(get_current_user)]) -> dict[str, Any]:
    ensure_company_access(user, data.company_id)
    with pool.connection() as conn:
        try:
            rows = conn.execute(
                """SELECT fa.asset_id, fa.asset_code, fa.asset_name,
                          ac.depreciation_expense_account_id, ac.accumulated_depreciation_account_id,
                          GREATEST(LEAST((fa.acquisition_cost-fa.residual_value)/fa.useful_life_months,
                               fa.acquisition_cost-fa.residual_value-fa.accumulated_depreciation),0)::NUMERIC(20,4) AS amount
                   FROM erp.fixed_assets fa JOIN erp.asset_categories ac ON ac.asset_category_id=fa.asset_category_id
                   WHERE fa.group_id=%s AND fa.company_id=%s AND fa.status='ACTIVE'
                     AND fa.placed_in_service_date<=%s
                     AND (fa.last_depreciation_date IS NULL OR fa.last_depreciation_date<%s)
                   FOR UPDATE OF fa""",
                (user["group_id"], data.company_id, data.depreciation_date, data.depreciation_date),
            ).fetchall()
            rows = [r for r in rows if Decimal(str(r["amount"])) > 0]
            if not rows:
                raise HTTPException(status_code=422, detail="لا توجد أصول مستحقة للإهلاك")
            entries = []
            for row in rows:
                entries.extend([
                    {"account_id": row["depreciation_expense_account_id"], "debit_amount": row["amount"], "credit_amount": 0, "description": f"إهلاك {row['asset_name']}"},
                    {"account_id": row["accumulated_depreciation_account_id"], "debit_amount": 0, "credit_amount": row["amount"], "description": f"مجمع إهلاك {row['asset_name']}"},
                ])
            voucher = create_voucher_db(
                conn, user=user, company_id=data.company_id,
                voucher_no=f"DEP-{data.depreciation_date.strftime('%Y%m%d')}",
                document_date=data.depreciation_date, posting_date=data.depreciation_date,
                description=f"إهلاك الأصول حتى {data.depreciation_date}", entries=entries,
                source_module="FA", external_reference=str(data.depreciation_date), post_immediately=True,
            )
            total = Decimal("0")
            for row in rows:
                total += Decimal(str(row["amount"]))
                conn.execute(
                    """INSERT INTO erp.asset_depreciation_entries
                           (group_id, company_id, asset_id, depreciation_date, amount, voucher_id)
                       VALUES (%s,%s,%s,%s,%s,%s)""",
                    (user["group_id"], data.company_id, row["asset_id"], data.depreciation_date, row["amount"], voucher["voucher_id"]),
                )
                conn.execute(
                    """UPDATE erp.fixed_assets
                       SET accumulated_depreciation=accumulated_depreciation+%s,
                           last_depreciation_date=%s WHERE asset_id=%s""",
                    (row["amount"], data.depreciation_date, row["asset_id"]),
                )
            audit(conn, user, "RUN", "DEPRECIATION", voucher["voucher_id"], data.company_id, {"assets": len(rows), "total": str(total)})
            conn.commit()
            return {"voucher_id": voucher["voucher_id"], "assets_count": len(rows), "total_amount": money(total)}
        except HTTPException:
            conn.rollback(); raise
        except Exception as exc:
            conn.rollback(); raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.get("/api/users")
def list_users(user: Annotated[dict[str, Any], Depends(get_current_user)]) -> list[dict[str, Any]]:
    require_group_admin(user)
    with pool.connection() as conn:
        return conn.execute(
            """SELECT u.user_id, u.full_name, u.email, u.role_code, u.is_group_admin,
                      u.company_id, c.company_name, u.is_active, u.last_login_at, u.created_at
               FROM erp.app_users u LEFT JOIN erp.companies c ON c.company_id=u.company_id
               WHERE u.group_id=%s ORDER BY u.created_at""",
            (user["group_id"],),
        ).fetchall()


@app.post("/api/users", status_code=201)
def create_user(data: UserCreate, user: Annotated[dict[str, Any], Depends(get_current_user)]) -> dict[str, Any]:
    require_group_admin(user)
    if data.company_id is None:
        raise HTTPException(status_code=422, detail="حدد شركة للمستخدم")
    with pool.connection() as conn:
        if data.company_id:
            company = conn.execute("SELECT 1 FROM erp.companies WHERE group_id=%s AND company_id=%s", (user["group_id"], data.company_id)).fetchone()
            if not company:
                raise HTTPException(status_code=422, detail="الشركة غير موجودة")
        row = conn.execute(
            """INSERT INTO erp.app_users
                   (group_id, company_id, email, password_hash, is_group_admin,
                    full_name, role_code)
               VALUES (%s,%s,%s,%s,FALSE,%s,%s)
               RETURNING user_id, full_name, email, role_code, company_id, is_active""",
            (user["group_id"], data.company_id, str(data.email).lower(), hash_password(data.password), data.full_name, data.role_code),
        ).fetchone()
        audit(conn, user, "CREATE", "USER", row["user_id"], data.company_id, {"role": data.role_code})
        conn.commit()
    return row


@app.patch("/api/users/{user_id}")
def update_user(user_id: UUID, data: UserUpdate, user: Annotated[dict[str, Any], Depends(get_current_user)]) -> dict[str, Any]:
    require_group_admin(user)
    with pool.connection() as conn:
        target = conn.execute(
            "SELECT user_id, is_group_admin FROM erp.app_users WHERE group_id=%s AND user_id=%s",
            (user["group_id"], user_id),
        ).fetchone()
        if not target:
            raise HTTPException(status_code=404, detail="المستخدم غير موجود")
        if target["is_group_admin"]:
            raise HTTPException(status_code=422, detail="لا يمكن تعديل حساب مدير المجموعة من هذه الشاشة")
        company = conn.execute(
            "SELECT 1 FROM erp.companies WHERE group_id=%s AND company_id=%s AND is_active=TRUE",
            (user["group_id"], data.company_id),
        ).fetchone()
        if not company:
            raise HTTPException(status_code=422, detail="الشركة غير موجودة")
        if data.password:
            row = conn.execute(
                """UPDATE erp.app_users
                   SET full_name=%s, role_code=%s, company_id=%s, password_hash=%s
                   WHERE group_id=%s AND user_id=%s
                   RETURNING user_id, full_name, email, role_code, company_id, is_active""",
                (data.full_name, data.role_code, data.company_id, hash_password(data.password),
                 user["group_id"], user_id),
            ).fetchone()
        else:
            row = conn.execute(
                """UPDATE erp.app_users
                   SET full_name=%s, role_code=%s, company_id=%s
                   WHERE group_id=%s AND user_id=%s
                   RETURNING user_id, full_name, email, role_code, company_id, is_active""",
                (data.full_name, data.role_code, data.company_id, user["group_id"], user_id),
            ).fetchone()
        audit(conn, user, "UPDATE", "USER", user_id, data.company_id, {"role": data.role_code})
        conn.commit()
    return row


@app.delete("/api/users/{user_id}")
def deactivate_user(user_id: UUID, user: Annotated[dict[str, Any], Depends(get_current_user)]) -> dict[str, str]:
    require_group_admin(user)
    if user_id == user["user_id"]:
        raise HTTPException(status_code=422, detail="لا يمكنك إيقاف حسابك الحالي")
    with pool.connection() as conn:
        target = conn.execute(
            "SELECT user_id, company_id, is_group_admin FROM erp.app_users WHERE group_id=%s AND user_id=%s AND is_active=TRUE",
            (user["group_id"], user_id),
        ).fetchone()
        if not target:
            raise HTTPException(status_code=404, detail="المستخدم غير موجود أو موقوف")
        if target["is_group_admin"]:
            raise HTTPException(status_code=422, detail="لا يمكن إيقاف حساب مدير المجموعة")
        conn.execute("UPDATE erp.app_users SET is_active=FALSE WHERE user_id=%s", (user_id,))
        audit(conn, user, "DEACTIVATE", "USER", user_id, target["company_id"])
        conn.commit()
    return {"status": "deactivated"}


@app.get("/api/trial-balance")
def trial_balance(user: Annotated[dict[str, Any], Depends(get_current_user)], company_id: UUID, as_of_date: date = Query(default_factory=date.today)) -> list[dict[str, Any]]:
    ensure_company_access(user, company_id)
    with pool.connection() as conn:
        return conn.execute(
            """SELECT a.local_account_code AS account_code, a.local_account_name AS account_name,
                      a.local_account_name_en AS account_name_en,
                      ga.account_class, SUM(e.debit_amount)::NUMERIC(20,4) AS total_debit,
                      SUM(e.credit_amount)::NUMERIC(20,4) AS total_credit,
                      SUM(e.debit_amount-e.credit_amount)::NUMERIC(20,4) AS net_balance
               FROM erp.journal_entries e
               JOIN erp.journal_vouchers v ON v.voucher_id=e.voucher_id AND v.company_id=e.company_id AND v.group_id=e.group_id
               JOIN erp.accounts a ON a.account_id=e.account_id AND a.company_id=e.company_id AND a.group_id=e.group_id
               JOIN erp.group_accounts ga ON ga.group_account_id=a.group_account_id
               WHERE e.group_id=%s AND e.company_id=%s AND v.status='POSTED' AND v.posting_date<=%s
               GROUP BY a.account_id, ga.account_class
               HAVING SUM(e.debit_amount)<>0 OR SUM(e.credit_amount)<>0 ORDER BY a.local_account_code""",
            (user["group_id"], company_id, as_of_date),
        ).fetchall()


@app.get("/api/general-ledger")
def general_ledger(user: Annotated[dict[str, Any], Depends(get_current_user)], company_id: UUID, account_id: UUID, date_from: date, date_to: date) -> list[dict[str, Any]]:
    ensure_company_access(user, company_id)
    with pool.connection() as conn:
        return conn.execute(
            """SELECT v.posting_date, v.voucher_no, v.source_module, v.description AS voucher_description,
                      e.line_no, e.entry_description, e.debit_amount, e.credit_amount,
                      SUM(e.debit_amount-e.credit_amount) OVER (ORDER BY v.posting_date, v.created_at, e.line_no) AS running_balance
               FROM erp.journal_entries e JOIN erp.journal_vouchers v ON v.voucher_id=e.voucher_id
               WHERE e.group_id=%s AND e.company_id=%s AND e.account_id=%s
                 AND v.status='POSTED' AND v.posting_date BETWEEN %s AND %s
               ORDER BY v.posting_date, v.created_at, e.line_no""",
            (user["group_id"], company_id, account_id, date_from, date_to),
        ).fetchall()


@app.get("/api/general-ledger-range")
def general_ledger_range(
    user: Annotated[dict[str, Any], Depends(get_current_user)],
    company_id: UUID,
    account_from: str,
    account_to: str,
    date_from: date,
    date_to: date,
) -> dict[str, Any]:
    ensure_company_access(user, company_id)
    if date_from > date_to:
        raise HTTPException(status_code=422, detail="تاريخ البداية يجب أن يسبق تاريخ النهاية")
    if account_from > account_to:
        raise HTTPException(status_code=422, detail="الحساب من يجب أن يسبق الحساب إلى")
    with pool.connection() as conn:
        company = conn.execute(
            """SELECT company_code, company_name, functional_currency
               FROM erp.companies WHERE group_id=%s AND company_id=%s""",
            (user["group_id"], company_id),
        ).fetchone()
        accounts = conn.execute(
            """SELECT account_id, local_account_code, local_account_name, local_account_name_en
               FROM erp.accounts
               WHERE group_id=%s AND company_id=%s AND is_active=TRUE
                 AND local_account_code BETWEEN %s AND %s
               ORDER BY local_account_code""",
            (user["group_id"], company_id, account_from, account_to),
        ).fetchall()
        if not accounts:
            raise HTTPException(status_code=404, detail="لا توجد حسابات في النطاق المحدد")
        account_ids = [row["account_id"] for row in accounts]
        openings = conn.execute(
            """SELECT e.account_id, COALESCE(SUM(e.debit_amount-e.credit_amount),0)::NUMERIC(20,4) AS opening_balance
               FROM erp.journal_entries e
               JOIN erp.journal_vouchers v ON v.voucher_id=e.voucher_id
               WHERE e.group_id=%s AND e.company_id=%s AND e.account_id=ANY(%s)
                 AND v.status='POSTED' AND v.posting_date<%s
               GROUP BY e.account_id""",
            (user["group_id"], company_id, account_ids, date_from),
        ).fetchall()
        entries = conn.execute(
            """SELECT e.account_id, v.posting_date, v.voucher_no, v.source_module,
                      COALESCE(e.entry_description, v.description, '') AS description,
                      e.debit_amount, e.credit_amount, v.created_at, e.line_no
               FROM erp.journal_entries e
               JOIN erp.journal_vouchers v ON v.voucher_id=e.voucher_id
               WHERE e.group_id=%s AND e.company_id=%s AND e.account_id=ANY(%s)
                 AND v.status='POSTED' AND v.posting_date BETWEEN %s AND %s
               ORDER BY e.account_id, v.posting_date, v.created_at, e.line_no""",
            (user["group_id"], company_id, account_ids, date_from, date_to),
        ).fetchall()

    opening_map = {row["account_id"]: money(row["opening_balance"]) for row in openings}
    entry_map: dict[UUID, list[dict[str, Any]]] = {}
    for entry in entries:
        entry_map.setdefault(entry["account_id"], []).append(entry)
    report_accounts: list[dict[str, Any]] = []
    total_debit = Decimal("0")
    total_credit = Decimal("0")
    for account in accounts:
        running = opening_map.get(account["account_id"], Decimal("0"))
        lines: list[dict[str, Any]] = []
        for entry in entry_map.get(account["account_id"], []):
            debit = money(entry["debit_amount"])
            credit = money(entry["credit_amount"])
            running = money(running + debit - credit)
            total_debit += debit
            total_credit += credit
            lines.append({
                "posting_date": entry["posting_date"],
                "voucher_no": entry["voucher_no"],
                "source_module": entry["source_module"],
                "description": entry["description"],
                "debit_amount": debit,
                "credit_amount": credit,
                "running_balance": running,
            })
        report_accounts.append({
            "account_id": account["account_id"],
            "account_code": account["local_account_code"],
            "account_name": account["local_account_name"],
            "account_name_en": account["local_account_name_en"],
            "opening_balance": opening_map.get(account["account_id"], Decimal("0")),
            "closing_balance": running,
            "entries": lines,
        })
    return {
        "company": company,
        "date_from": date_from,
        "date_to": date_to,
        "account_from": account_from,
        "account_to": account_to,
        "accounts": report_accounts,
        "total_debit": money(total_debit),
        "total_credit": money(total_credit),
    }


@app.get("/api/income-statement")
def income_statement(user: Annotated[dict[str, Any], Depends(get_current_user)], company_id: UUID, date_from: date, date_to: date) -> dict[str, Any]:
    ensure_company_access(user, company_id)
    with pool.connection() as conn:
        rows = conn.execute(
            """SELECT ga.account_class, a.local_account_code AS account_code, a.local_account_name AS account_name,
                      a.local_account_name_en AS account_name_en,
                      CASE WHEN ga.account_class='REVENUE' THEN SUM(e.credit_amount-e.debit_amount)
                           ELSE SUM(e.debit_amount-e.credit_amount) END::NUMERIC(20,4) AS amount
               FROM erp.journal_entries e JOIN erp.journal_vouchers v ON v.voucher_id=e.voucher_id
               JOIN erp.accounts a ON a.account_id=e.account_id
               JOIN erp.group_accounts ga ON ga.group_account_id=a.group_account_id
               WHERE e.group_id=%s AND e.company_id=%s AND v.status='POSTED'
                 AND v.posting_date BETWEEN %s AND %s AND ga.account_class IN ('REVENUE','EXPENSE')
               GROUP BY ga.account_class, a.account_id ORDER BY ga.account_class DESC, a.local_account_code""",
            (user["group_id"], company_id, date_from, date_to),
        ).fetchall()
    revenues = sum(Decimal(str(r["amount"])) for r in rows if r["account_class"] == "REVENUE")
    expenses = sum(Decimal(str(r["amount"])) for r in rows if r["account_class"] == "EXPENSE")
    return {"rows": rows, "total_revenue": money(revenues), "total_expense": money(expenses), "net_profit": money(revenues-expenses)}


@app.get("/api/balance-sheet")
def balance_sheet(user: Annotated[dict[str, Any], Depends(get_current_user)], company_id: UUID, as_of_date: date) -> dict[str, Any]:
    ensure_company_access(user, company_id)
    with pool.connection() as conn:
        rows = conn.execute(
            """SELECT ga.account_class, a.local_account_code AS account_code, a.local_account_name AS account_name,
                      a.local_account_name_en AS account_name_en,
                      CASE WHEN ga.account_class='ASSET' THEN SUM(e.debit_amount-e.credit_amount)
                           ELSE SUM(e.credit_amount-e.debit_amount) END::NUMERIC(20,4) AS amount
               FROM erp.journal_entries e JOIN erp.journal_vouchers v ON v.voucher_id=e.voucher_id
               JOIN erp.accounts a ON a.account_id=e.account_id
               JOIN erp.group_accounts ga ON ga.group_account_id=a.group_account_id
               WHERE e.group_id=%s AND e.company_id=%s AND v.status='POSTED'
                 AND v.posting_date<=%s AND ga.account_class IN ('ASSET','LIABILITY','EQUITY')
               GROUP BY ga.account_class, a.account_id ORDER BY ga.account_class, a.local_account_code""",
            (user["group_id"], company_id, as_of_date),
        ).fetchall()
        current_result = conn.execute(
            """SELECT COALESCE(SUM(
                         CASE WHEN ga.account_class='REVENUE'
                              THEN e.credit_amount-e.debit_amount
                              ELSE e.debit_amount-e.credit_amount END
                       ),0)::NUMERIC(20,4) AS amount
               FROM erp.journal_entries e
               JOIN erp.journal_vouchers v ON v.voucher_id=e.voucher_id
               JOIN erp.accounts a ON a.account_id=e.account_id
               JOIN erp.group_accounts ga ON ga.group_account_id=a.group_account_id
               WHERE e.group_id=%s AND e.company_id=%s AND v.status='POSTED'
                 AND v.posting_date<=%s AND ga.account_class IN ('REVENUE','EXPENSE')""",
            (user["group_id"], company_id, as_of_date),
        ).fetchone()["amount"]
    rows = list(rows)
    if money(current_result) != 0:
        rows.append({
            "account_class": "EQUITY",
            "account_code": "CURRENT_RESULT",
            "account_name": "نتيجة الفترة — أرباح / خسائر",
            "account_name_en": "Profit or Loss for the Period",
            "amount": money(current_result),
        })
    totals = {c: money(sum(Decimal(str(r["amount"])) for r in rows if r["account_class"] == c)) for c in ("ASSET", "LIABILITY", "EQUITY")}
    return {"rows": rows, "totals": totals, "difference": money(totals["ASSET"]-totals["LIABILITY"]-totals["EQUITY"])}


@app.get("/api/consolidated-income-statement")
def consolidated_income_statement(
    user: Annotated[dict[str, Any], Depends(get_current_user)],
    date_from: date,
    date_to: date,
) -> dict[str, Any]:
    require_group_admin(user)
    if date_from > date_to:
        raise HTTPException(status_code=422, detail="تاريخ البداية يجب أن يسبق تاريخ النهاية")
    with pool.connection() as conn:
        rows = conn.execute(
            """SELECT ga.account_class, ga.account_code, ga.account_name, ga.account_name_en,
                      CASE WHEN ga.account_class='REVENUE' THEN SUM(e.credit_amount-e.debit_amount)
                           ELSE SUM(e.debit_amount-e.credit_amount) END::NUMERIC(20,4) AS amount
               FROM erp.journal_entries e
               JOIN erp.journal_vouchers v ON v.voucher_id=e.voucher_id
               JOIN erp.accounts a ON a.account_id=e.account_id
               JOIN erp.group_accounts ga ON ga.group_account_id=a.group_account_id
               JOIN erp.companies c ON c.company_id=e.company_id
               WHERE e.group_id=%s AND v.status='POSTED'
                 AND v.posting_date BETWEEN %s AND %s
                 AND ga.account_class IN ('REVENUE','EXPENSE')
                 AND c.is_active IN (TRUE,FALSE)
               GROUP BY ga.group_account_id
               HAVING SUM(e.debit_amount)<>0 OR SUM(e.credit_amount)<>0
               ORDER BY ga.account_class DESC, ga.account_code""",
            (user["group_id"], date_from, date_to),
        ).fetchall()
    revenues = sum(Decimal(str(row["amount"])) for row in rows if row["account_class"] == "REVENUE")
    expenses = sum(Decimal(str(row["amount"])) for row in rows if row["account_class"] == "EXPENSE")
    return {
        "rows": rows,
        "total_revenue": money(revenues),
        "total_expense": money(expenses),
        "net_profit": money(revenues - expenses),
        "date_from": date_from,
        "date_to": date_to,
    }


@app.get("/api/consolidated-balance-sheet")
def consolidated_balance_sheet(
    user: Annotated[dict[str, Any], Depends(get_current_user)],
    as_of_date: date,
) -> dict[str, Any]:
    require_group_admin(user)
    with pool.connection() as conn:
        rows = conn.execute(
            """SELECT ga.account_class, ga.account_code, ga.account_name, ga.account_name_en,
                      CASE WHEN ga.account_class='ASSET' THEN SUM(e.debit_amount-e.credit_amount)
                           ELSE SUM(e.credit_amount-e.debit_amount) END::NUMERIC(20,4) AS amount
               FROM erp.journal_entries e
               JOIN erp.journal_vouchers v ON v.voucher_id=e.voucher_id
               JOIN erp.accounts a ON a.account_id=e.account_id
               JOIN erp.group_accounts ga ON ga.group_account_id=a.group_account_id
               JOIN erp.companies c ON c.company_id=e.company_id
               WHERE e.group_id=%s AND v.status='POSTED' AND v.posting_date<=%s
                 AND ga.account_class IN ('ASSET','LIABILITY','EQUITY')
                 AND c.is_active IN (TRUE,FALSE)
               GROUP BY ga.group_account_id
               HAVING SUM(e.debit_amount)<>0 OR SUM(e.credit_amount)<>0
               ORDER BY ga.account_class, ga.account_code""",
            (user["group_id"], as_of_date),
        ).fetchall()
        current_result = conn.execute(
            """SELECT COALESCE(SUM(
                         CASE WHEN ga.account_class='REVENUE'
                              THEN e.credit_amount-e.debit_amount
                              ELSE e.debit_amount-e.credit_amount END
                       ),0)::NUMERIC(20,4) AS amount
               FROM erp.journal_entries e
               JOIN erp.journal_vouchers v ON v.voucher_id=e.voucher_id
               JOIN erp.accounts a ON a.account_id=e.account_id
               JOIN erp.group_accounts ga ON ga.group_account_id=a.group_account_id
               WHERE e.group_id=%s AND v.status='POSTED' AND v.posting_date<=%s
                 AND ga.account_class IN ('REVENUE','EXPENSE')""",
            (user["group_id"], as_of_date),
        ).fetchone()["amount"]
    rows = list(rows)
    if money(current_result) != 0:
        rows.append({
            "account_class": "EQUITY",
            "account_code": "CURRENT_RESULT",
            "account_name": "نتيجة الفترة المجمعة — أرباح / خسائر",
            "account_name_en": "Consolidated Profit or Loss for the Period",
            "amount": money(current_result),
        })
    totals = {
        account_class: money(sum(
            Decimal(str(row["amount"]))
            for row in rows if row["account_class"] == account_class
        ))
        for account_class in ("ASSET", "LIABILITY", "EQUITY")
    }
    return {
        "rows": rows,
        "totals": totals,
        "difference": money(totals["ASSET"] - totals["LIABILITY"] - totals["EQUITY"]),
        "as_of_date": as_of_date,
    }


def previous_year_date(value: date) -> date:
    return date(value.year - 1, value.month, min(value.day, calendar.monthrange(value.year - 1, value.month)[1]))


@app.get("/api/ifrs-statement")
def ifrs_statement(
    user: Annotated[dict[str, Any], Depends(get_current_user)],
    statement: Literal["INCOME", "POSITION", "EQUITY", "CASH_FLOW"],
    date_from: date,
    date_to: date,
    company_id: UUID | None = None,
    comparative_from: date | None = None,
    comparative_to: date | None = None,
) -> dict[str, Any]:
    if date_from > date_to:
        raise HTTPException(status_code=422, detail="تاريخ البداية يجب أن يسبق تاريخ النهاية")
    if company_id:
        ensure_company_access(user, company_id)
    else:
        require_group_admin(user)
    comparative_from = comparative_from or previous_year_date(date_from)
    comparative_to = comparative_to or previous_year_date(date_to)
    if comparative_from > comparative_to:
        raise HTTPException(status_code=422, detail="فترة المقارنة غير صحيحة")

    with pool.connection() as conn:
        company = None
        if company_id:
            company = conn.execute(
                """SELECT company_code, company_name, functional_currency
                   FROM erp.companies WHERE group_id=%s AND company_id=%s""",
                (user["group_id"], company_id),
            ).fetchone()

        def account_amounts(start: date, end: date, as_of: bool = False) -> list[dict[str, Any]]:
            classes = ("ASSET", "LIABILITY", "EQUITY") if as_of else ("REVENUE", "EXPENSE")
            date_sql = "v.posting_date<=%s" if as_of else "v.posting_date BETWEEN %s AND %s"
            params: list[Any] = [user["group_id"], company_id, company_id]
            params.extend([end] if as_of else [start, end])
            params.append(list(classes))
            return conn.execute(
                f"""SELECT ga.account_code, ga.account_name, ga.account_name_en, ga.account_class,
                           SUM(e.debit_amount-e.credit_amount)::NUMERIC(20,4) AS net
                    FROM erp.journal_entries e
                    JOIN erp.journal_vouchers v ON v.voucher_id=e.voucher_id
                    JOIN erp.accounts a ON a.account_id=e.account_id
                    JOIN erp.group_accounts ga ON ga.group_account_id=a.group_account_id
                    WHERE e.group_id=%s AND (%s::uuid IS NULL OR e.company_id=%s)
                      AND v.status='POSTED' AND {date_sql}
                      AND ga.account_class=ANY(%s)
                    GROUP BY ga.group_account_id
                    HAVING SUM(e.debit_amount)<>0 OR SUM(e.credit_amount)<>0
                    ORDER BY ga.account_code""",
                params,
            ).fetchall()

        def period_profit(start: date, end: date) -> Decimal:
            row = conn.execute(
                """SELECT COALESCE(SUM(
                             CASE WHEN ga.account_class='REVENUE'
                                  THEN e.credit_amount-e.debit_amount
                                  ELSE e.debit_amount-e.credit_amount END
                           ),0)::NUMERIC(20,4) AS amount
                   FROM erp.journal_entries e
                   JOIN erp.journal_vouchers v ON v.voucher_id=e.voucher_id
                   JOIN erp.accounts a ON a.account_id=e.account_id
                   JOIN erp.group_accounts ga ON ga.group_account_id=a.group_account_id
                   WHERE e.group_id=%s AND (%s::uuid IS NULL OR e.company_id=%s)
                     AND v.status='POSTED' AND v.posting_date BETWEEN %s AND %s
                     AND ga.account_class IN ('REVENUE','EXPENSE')""",
                (user["group_id"], company_id, company_id, start, end),
            ).fetchone()
            return money(row["amount"])

        def cash_flow(start: date, end: date) -> dict[str, Decimal]:
            rows = conn.execute(
                """WITH voucher_cash AS (
                     SELECT v.voucher_id,
                            SUM(CASE WHEN ga.account_code LIKE '111%%' OR ga.account_code LIKE '112%%'
                                     THEN e.debit_amount-e.credit_amount ELSE 0 END) AS cash_change,
                            BOOL_OR(ga.account_code LIKE '12%%') AS has_investing,
                            BOOL_OR(ga.account_code LIKE '22%%' OR ga.account_class='EQUITY') AS has_financing
                     FROM erp.journal_vouchers v
                     JOIN erp.journal_entries e ON e.voucher_id=v.voucher_id
                     JOIN erp.accounts a ON a.account_id=e.account_id
                     JOIN erp.group_accounts ga ON ga.group_account_id=a.group_account_id
                     WHERE v.group_id=%s AND (%s::uuid IS NULL OR v.company_id=%s)
                       AND v.status='POSTED' AND v.posting_date BETWEEN %s AND %s
                     GROUP BY v.voucher_id
                   )
                   SELECT CASE WHEN has_investing THEN 'INVESTING'
                               WHEN has_financing THEN 'FINANCING'
                               ELSE 'OPERATING' END AS section,
                          COALESCE(SUM(cash_change),0)::NUMERIC(20,4) AS amount
                   FROM voucher_cash WHERE cash_change<>0 GROUP BY 1""",
                (user["group_id"], company_id, company_id, start, end),
            ).fetchall()
            result = {"OPERATING": Decimal("0"), "INVESTING": Decimal("0"), "FINANCING": Decimal("0")}
            for row in rows:
                result[row["section"]] = money(row["amount"])
            return result

        current_source: list[dict[str, Any]] = []
        comparative_source: list[dict[str, Any]] = []
        current_profit = Decimal("0")
        comparative_profit = Decimal("0")
        current_cash: dict[str, Decimal] = {}
        comparative_cash: dict[str, Decimal] = {}
        if statement in ("INCOME", "POSITION"):
            current_source = account_amounts(date_from, date_to, as_of=statement == "POSITION")
            comparative_source = account_amounts(comparative_from, comparative_to, as_of=statement == "POSITION")
            if statement == "POSITION":
                current_profit = period_profit(date(date_to.year, 1, 1), date_to)
                comparative_profit = period_profit(date(comparative_to.year, 1, 1), comparative_to)
        elif statement == "EQUITY":
            current_source = account_amounts(date_from, date_to, as_of=True)
            comparative_source = account_amounts(comparative_from, comparative_to, as_of=True)
            current_profit = period_profit(date_from, date_to)
            comparative_profit = period_profit(comparative_from, comparative_to)
        else:
            current_cash = cash_flow(date_from, date_to)
            comparative_cash = cash_flow(comparative_from, comparative_to)

    rows: list[dict[str, Any]] = []
    if statement in ("INCOME", "POSITION"):
        current = {row["account_code"]: row for row in current_source}
        comparative = {row["account_code"]: row for row in comparative_source}
        for code in sorted(set(current) | set(comparative)):
            base = current.get(code) or comparative[code]
            account_class = base["account_class"]
            current_net = Decimal(str(current.get(code, {}).get("net", 0)))
            comparative_net = Decimal(str(comparative.get(code, {}).get("net", 0)))
            sign = Decimal("-1") if account_class in ("LIABILITY", "EQUITY", "REVENUE") else Decimal("1")
            section = account_class
            if statement == "POSITION":
                if account_class == "ASSET":
                    section = "CURRENT_ASSET" if code.startswith("11") else "NON_CURRENT_ASSET"
                elif account_class == "LIABILITY":
                    section = "CURRENT_LIABILITY" if code.startswith("21") else "NON_CURRENT_LIABILITY"
            rows.append({
                "account_code": code,
                "account_name": base["account_name"],
                "account_name_en": base["account_name_en"],
                "section": section,
                "current_amount": money(current_net * sign),
                "comparative_amount": money(comparative_net * sign),
            })
        if statement == "POSITION":
            rows.append({
                "account_code": "CURRENT_RESULT",
                "account_name": "نتيجة الفترة — أرباح / خسائر",
                "account_name_en": "Profit or Loss for the Period",
                "section": "EQUITY",
                "current_amount": current_profit,
                "comparative_amount": comparative_profit,
            })
    elif statement == "EQUITY":
        current = {row["account_code"]: row for row in current_source if row["account_class"] == "EQUITY"}
        comparative = {row["account_code"]: row for row in comparative_source if row["account_class"] == "EQUITY"}
        for code in sorted(set(current) | set(comparative)):
            base = current.get(code) or comparative[code]
            rows.append({
                "account_code": code,
                "account_name": base["account_name"],
                "account_name_en": base["account_name_en"],
                "section": "EQUITY",
                "current_amount": money(-Decimal(str(current.get(code, {}).get("net", 0)))),
                "comparative_amount": money(-Decimal(str(comparative.get(code, {}).get("net", 0)))),
            })
        rows.append({
            "account_code": "PERIOD_RESULT",
            "account_name": "صافي ربح / خسارة الفترة",
            "account_name_en": "Net Profit or Loss for the Period",
            "section": "EQUITY",
            "current_amount": current_profit,
            "comparative_amount": comparative_profit,
        })
    else:
        names = {
            "OPERATING": "صافي التدفقات النقدية من الأنشطة التشغيلية",
            "INVESTING": "صافي التدفقات النقدية من الأنشطة الاستثمارية",
            "FINANCING": "صافي التدفقات النقدية من الأنشطة التمويلية",
        }
        names_en = {
            "OPERATING": "Net Cash Flows from Operating Activities",
            "INVESTING": "Net Cash Flows from Investing Activities",
            "FINANCING": "Net Cash Flows from Financing Activities",
        }
        for section in ("OPERATING", "INVESTING", "FINANCING"):
            rows.append({
                "account_code": section,
                "account_name": names[section],
                "account_name_en": names_en[section],
                "section": section,
                "current_amount": current_cash[section],
                "comparative_amount": comparative_cash[section],
            })
        rows.append({
            "account_code": "NET_CHANGE",
            "account_name": "صافي التغير في النقدية وما في حكمها",
            "account_name_en": "Net Change in Cash and Cash Equivalents",
            "section": "TOTAL",
            "current_amount": money(sum(current_cash.values())),
            "comparative_amount": money(sum(comparative_cash.values())),
        })

    totals: dict[str, dict[str, Decimal]] = {}
    for row in rows:
        section = row["section"]
        totals.setdefault(section, {"current": Decimal("0"), "comparative": Decimal("0")})
        totals[section]["current"] = money(totals[section]["current"] + Decimal(str(row["current_amount"])))
        totals[section]["comparative"] = money(totals[section]["comparative"] + Decimal(str(row["comparative_amount"])))
    result = {
        "statement": statement,
        "consolidated": company_id is None,
        "company": company,
        "currency": company["functional_currency"] if company else "EGP",
        "date_from": date_from,
        "date_to": date_to,
        "comparative_from": comparative_from,
        "comparative_to": comparative_to,
        "rows": rows,
        "totals": totals,
        "framework": "IAS 1 / IAS 7 presentation framework; IFRS 18 readiness",
        "ifrs_note": "Generated from posted ledger balances. Final IFRS compliance depends on account mapping, estimates, disclosures and professional review.",
    }
    return result


@app.get("/api/consolidated-trial-balance")
def consolidated_trial_balance(user: Annotated[dict[str, Any], Depends(get_current_user)], as_of_date: date = Query(default_factory=date.today)) -> list[dict[str, Any]]:
    require_group_admin(user)
    with pool.connection() as conn:
        rows = conn.execute(
            """SELECT ga.account_code, ga.account_name, ga.account_name_en, c.company_code,
                      SUM(e.debit_amount-e.credit_amount)::NUMERIC(20,4) AS net_balance
               FROM erp.journal_entries e
               JOIN erp.journal_vouchers v ON v.voucher_id=e.voucher_id
               JOIN erp.accounts a ON a.account_id=e.account_id
               JOIN erp.group_accounts ga ON ga.group_account_id=a.group_account_id
               JOIN erp.companies c ON c.company_id=e.company_id
               WHERE e.group_id=%s AND v.status='POSTED' AND v.posting_date<=%s
               GROUP BY ga.group_account_id, c.company_code ORDER BY ga.account_code, c.company_code""",
            (user["group_id"], as_of_date),
        ).fetchall()
    pivot: dict[str, dict[str, Any]] = {}
    for row in rows:
        item = pivot.setdefault(row["account_code"], {"account_code": row["account_code"], "account_name": row["account_name"], "account_name_en": row["account_name_en"], "companies": {}, "consolidated_net": 0})
        amount = float(row["net_balance"])
        item["companies"][row["company_code"]] = amount
        item["consolidated_net"] += amount
    return list(pivot.values())


@app.get("/api/audit-log")
def audit_log(user: Annotated[dict[str, Any], Depends(get_current_user)], limit: int = Query(100, ge=1, le=500)) -> list[dict[str, Any]]:
    require_group_admin(user)
    with pool.connection() as conn:
        return conn.execute(
            """SELECT al.audit_id, al.created_at, al.action, al.entity_type, al.entity_id,
                      al.company_id, al.details, u.full_name, u.email, c.company_name
               FROM erp.audit_log al LEFT JOIN erp.app_users u ON u.user_id=al.user_id
               LEFT JOIN erp.companies c ON c.company_id=al.company_id
               WHERE al.group_id=%s ORDER BY al.created_at DESC LIMIT %s""",
            (user["group_id"], limit),
        ).fetchall()
